import io
import os
import calendar
from datetime import datetime, date, timezone, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file

JST = timezone(timedelta(hours=9))

def now_jst():
    return datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")

def today_jst():
    return datetime.now(JST).date()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "shift-secret-key")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "saito")

DATABASE_URL = os.environ.get("DATABASE_URL")


def get_db():
    if DATABASE_URL:
        import psycopg2, psycopg2.extras
        url = DATABASE_URL.replace("postgres://", "postgresql://", 1)
        conn = psycopg2.connect(url)
        conn.cursor_factory = psycopg2.extras.RealDictCursor
        return conn
    else:
        import sqlite3
        db_path = os.path.join(os.path.dirname(__file__), "shift.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn

def ph():
    return "%s" if DATABASE_URL else "?"


def execute(sql, params=()):
    p = ph()
    sql = sql.replace("?", p)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        conn.commit()
        return cur
    finally:
        conn.close()


def query(sql, params=()):
    p = ph()
    sql = sql.replace("?", p)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


def init_db():
    serial = "SERIAL" if DATABASE_URL else "INTEGER"
    ai     = "" if DATABASE_URL else "AUTOINCREMENT"
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS shift_staff (
                id               {serial} PRIMARY KEY {ai},
                name             TEXT NOT NULL UNIQUE,
                color            TEXT NOT NULL DEFAULT '#c87941',
                allowed_slots    TEXT NOT NULL DEFAULT '',
                max_days         INTEGER NOT NULL DEFAULT 0,
                wed_ok           INTEGER NOT NULL DEFAULT 0,
                sun_ok           INTEGER NOT NULL DEFAULT 0,
                sun_a_exclusive  INTEGER NOT NULL DEFAULT 0,
                same_day_ng      TEXT NOT NULL DEFAULT '',
                days_off_str     TEXT NOT NULL DEFAULT ''
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS shift_records (
                id         {serial} PRIMARY KEY {ai},
                staff_id   INTEGER NOT NULL,
                shift_date TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time   TEXT NOT NULL,
                slot_label TEXT NOT NULL DEFAULT '',
                memo       TEXT DEFAULT ''
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS shift_requests (
                id           {serial} PRIMARY KEY {ai},
                staff_name   TEXT NOT NULL,
                month        TEXT NOT NULL,
                days_off_str TEXT NOT NULL DEFAULT '',
                submitted_at TEXT NOT NULL,
                UNIQUE(staff_name, month)
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS shift_day_rules (
                dow         INTEGER NOT NULL UNIQUE,
                is_closed   INTEGER NOT NULL DEFAULT 0,
                slot_a      INTEGER NOT NULL DEFAULT 0,
                slot_b      INTEGER NOT NULL DEFAULT 0,
                slot_c      INTEGER NOT NULL DEFAULT 0,
                slot_d      INTEGER NOT NULL DEFAULT 0,
                slot_m      INTEGER NOT NULL DEFAULT 0,
                slot_shikomi INTEGER NOT NULL DEFAULT 0
            )
        """)
        conn.commit()
        # デフォルトの曜日設定を挿入（存在しない場合のみ）
        defaults = [
            # dow, closed, A, B, C, D, M, 仕込み
            (0, 0, 1, 1, 1, 0, 0, 0),  # 月
            (1, 0, 0, 0, 0, 0, 0, 1),  # 火
            (2, 0, 0, 0, 0, 0, 0, 1),  # 水
            (3, 0, 1, 1, 1, 0, 0, 0),  # 木
            (4, 0, 1, 1, 1, 0, 0, 0),  # 金
            (5, 1, 0, 0, 0, 0, 0, 0),  # 土（休業）
            (6, 0, 1, 1, 1, 1, 0, 0),  # 日
        ]
        for row in defaults:
            try:
                cur.execute(
                    "INSERT INTO shift_day_rules (dow,is_closed,slot_a,slot_b,slot_c,slot_d,slot_m,slot_shikomi) "
                    "VALUES (?,?,?,?,?,?,?,?)".replace("?", "%s" if DATABASE_URL else "?"),
                    row
                )
                conn.commit()
            except Exception:
                conn.rollback()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS shift_config (
                id             {serial} PRIMARY KEY {ai},
                month          TEXT NOT NULL UNIQUE,
                marche_dates   TEXT NOT NULL DEFAULT '',
                slot_a_label   TEXT NOT NULL DEFAULT '6:30〜製造・品出し',
                slot_b_label   TEXT NOT NULL DEFAULT '9:00〜レジ',
                slot_c_label   TEXT NOT NULL DEFAULT '9:00〜レジ補助',
                slot_d_label   TEXT NOT NULL DEFAULT '10:00〜品出し(日曜)',
                slot_m_label   TEXT NOT NULL DEFAULT '7:00〜マルシェ',
                marche_m_count INTEGER NOT NULL DEFAULT 2
            )
        """)
        conn.commit()
        # マイグレーション
        for col, defval in [
            ("allowed_slots",   "TEXT NOT NULL DEFAULT ''"),
            ("max_days",        "INTEGER NOT NULL DEFAULT 0"),
            ("wed_ok",          "INTEGER NOT NULL DEFAULT 0"),
            ("sun_ok",          "INTEGER NOT NULL DEFAULT 0"),
            ("sun_a_exclusive", "INTEGER NOT NULL DEFAULT 0"),
            ("same_day_ng",     "TEXT NOT NULL DEFAULT ''"),
            ("days_off_str",    "TEXT NOT NULL DEFAULT ''"),
            ("min_sun_days",    "INTEGER NOT NULL DEFAULT 0"),
            ("sun_only",        "INTEGER NOT NULL DEFAULT 0"),
            ("wants_more",      "INTEGER NOT NULL DEFAULT 0"),
        ]:
            try:
                cur.execute(f"ALTER TABLE shift_staff ADD COLUMN {col} {defval}")
                conn.commit()
            except Exception:
                conn.rollback()
        for col, defval in [("slot_label", "TEXT NOT NULL DEFAULT ''")]:
            try:
                cur.execute(f"ALTER TABLE shift_records ADD COLUMN {col} {defval}")
                conn.commit()
            except Exception:
                conn.rollback()
    finally:
        conn.close()


# ── Auth ──────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect(url_for("admin"))
        error = "パスワードが違います"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


def require_login():
    if not session.get("admin"):
        return redirect(url_for("index"))
    return None


# ── 公開トップページ（カレンダー＋希望提出＋ログイン） ──────────────

@app.route("/", methods=["GET", "POST"])
def index():
    today = today_jst()
    year  = int(request.args.get("year",  today.year))
    month = int(request.args.get("month", today.month))

    # ── カレンダーデータ ──
    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])
    cal       = calendar.monthcalendar(year, month)
    staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
    shifts_raw = query(
        "SELECT s.*, st.name as staff_name, st.color FROM shift_records s "
        "JOIN shift_staff st ON s.staff_id = st.id "
        "WHERE s.shift_date BETWEEN ? AND ?",
        (first_day.isoformat(), last_day.isoformat())
    )
    shift_map = {}
    for sh in shifts_raw:
        shift_map.setdefault(sh["shift_date"], []).append(sh)
    for d in shift_map:
        shift_map[d].sort(key=lambda x: x["start_time"])

    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    # ── シフト希望提出 ──
    months = []
    for delta in range(0, 3):
        if today.month + delta <= 12:
            months.append(date(today.year, today.month + delta, 1))
        else:
            months.append(date(today.year + 1, (today.month + delta) - 12, 1))

    sel_month_str = request.args.get("req_month", months[0].strftime("%Y-%m"))
    try:
        sel_year, sel_month = map(int, sel_month_str.split("-"))
    except Exception:
        sel_year, sel_month = today.year, today.month
    days_in_month_req = calendar.monthrange(sel_year, sel_month)[1]

    req_msg = req_error = None
    login_error = None

    if request.method == "POST":
        action = request.form.get("action")

        # シフト希望送信
        if action == "request":
            staff_name   = request.form.get("staff_name", "").strip()
            month_str    = request.form.get("req_month_val", "").strip()
            days_off     = sorted(set(request.form.getlist("days_off")), key=int)
            days_off_str = ",".join(days_off)
            if not staff_name or not month_str:
                req_error = "名前と対象月を選択してください"
            else:
                existing = query(
                    "SELECT id FROM shift_requests WHERE staff_name=? AND month=?",
                    (staff_name, month_str)
                )
                if existing:
                    execute(
                        "UPDATE shift_requests SET days_off_str=?, submitted_at=? WHERE staff_name=? AND month=?",
                        (days_off_str, now_jst(), staff_name, month_str)
                    )
                else:
                    execute(
                        "INSERT INTO shift_requests (staff_name, month, days_off_str, submitted_at) VALUES (?,?,?,?)",
                        (staff_name, month_str, days_off_str, now_jst())
                    )
                execute(
                    "UPDATE shift_staff SET days_off_str=? WHERE name=?",
                    (days_off_str, staff_name)
                )
                req_msg = f"{staff_name} さんの {month_str} の希望を受け付けました ✅"

        # 管理者ログイン
        elif action == "login":
            if request.form.get("password") == ADMIN_PASSWORD:
                session["admin"] = True
                return redirect(url_for("admin"))
            else:
                login_error = "パスワードが違います"

    return render_template(
        "index.html",
        # カレンダー
        year=year, month=month, cal=cal,
        shift_map=shift_map, staff_rows=staff_rows,
        today=today,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        month_name=f"{year}年{month}月",
        # 希望提出
        months=months, sel_month_str=sel_month_str,
        sel_year=sel_year, sel_month=sel_month,
        days_in_month_req=days_in_month_req,
        req_msg=req_msg, req_error=req_error,
        # ログイン
        login_error=login_error,
        is_admin=session.get("admin", False),
    )


# ── シフト希望削除 ───────────────────────────────────────────────────────

@app.route("/request/delete/<int:req_id>", methods=["POST"])
def request_delete(req_id):
    r = require_login()
    if r:
        return r
    execute("DELETE FROM shift_requests WHERE id=?", (req_id,))
    return redirect(url_for("generate"))


# ── シフト希望ページ ────────────────────────────────────────────────────

@app.route("/request", methods=["GET", "POST"])
def shift_request():
    today = today_jst()
    months = []
    for delta in range(0, 6):
        m = today.month + delta
        y = today.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        months.append(date(y, m, 1))

    sel_month_str = request.args.get("month", months[0].strftime("%Y-%m"))
    try:
        sel_year, sel_month = map(int, sel_month_str.split("-"))
    except Exception:
        sel_year, sel_month = today.year, today.month
    days_in_month = calendar.monthrange(sel_year, sel_month)[1]
    cal_offset = date(sel_year, sel_month, 1).weekday()  # 0=月, 6=日

    msg = error = None
    if request.method == "POST":
        staff_name   = request.form.get("staff_name", "").strip()
        month_str    = request.form.get("month", "").strip()
        days_off     = sorted(set(request.form.getlist("days_off")), key=int)
        days_off_str = ",".join(days_off)
        if not staff_name or not month_str:
            error = "名前と対象月を選択してください"
        else:
            existing = query(
                "SELECT id FROM shift_requests WHERE staff_name=? AND month=?",
                (staff_name, month_str)
            )
            if existing:
                execute(
                    "UPDATE shift_requests SET days_off_str=?, submitted_at=? WHERE staff_name=? AND month=?",
                    (days_off_str, now_jst(), staff_name, month_str)
                )
            else:
                execute(
                    "INSERT INTO shift_requests (staff_name, month, days_off_str, submitted_at) VALUES (?,?,?,?)",
                    (staff_name, month_str, days_off_str, now_jst())
                )
            execute(
                "UPDATE shift_staff SET days_off_str=? WHERE name=?",
                (days_off_str, staff_name)
            )
            msg = f"{staff_name} さんの {month_str} の希望を受け付けました"

    staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
    return render_template(
        "request.html",
        staff_rows=staff_rows, today=today,
        months=months, sel_month_str=sel_month_str,
        sel_year=sel_year, sel_month=sel_month,
        days_in_month=days_in_month,
        cal_offset=cal_offset,
        msg=msg, error=error,
    )


# ── 管理者画面 ────────────────────────────────────────────────────────

@app.route("/admin")
def admin():
    if not session.get("admin"):
        return redirect(url_for("index"))
    today = today_jst()
    year  = int(request.args.get("year",  today.year))
    month = int(request.args.get("month", today.month))
    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])
    cal       = calendar.monthcalendar(year, month)
    staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
    shifts_raw = query(
        "SELECT s.*, st.name as staff_name, st.color FROM shift_records s "
        "JOIN shift_staff st ON s.staff_id = st.id "
        "WHERE s.shift_date BETWEEN ? AND ?",
        (first_day.isoformat(), last_day.isoformat())
    )
    shift_map = {}
    for sh in shifts_raw:
        shift_map.setdefault(sh["shift_date"], []).append(sh)
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    return render_template(
        "admin_calendar.html",
        year=year, month=month, cal=cal,
        shift_map=shift_map, staff_rows=staff_rows,
        today=today,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        month_name=f"{year}年{month}月",
    )


# ── シフト追加 ────────────────────────────────────────────────────────

@app.route("/shift/add", methods=["POST"])
def shift_add():
    r = require_login()
    if r:
        return r

    staff_id   = request.form.get("staff_id", "").strip()
    shift_date = request.form.get("shift_date", "").strip()
    start_time = request.form.get("start_time", "").strip()
    end_time   = request.form.get("end_time", "").strip()
    memo       = request.form.get("memo", "").strip()

    if not all([staff_id, shift_date, start_time, end_time]):
        return redirect(request.referrer or url_for("index"))

    execute(
        "INSERT INTO shift_records (staff_id, shift_date, start_time, end_time, memo) VALUES (?, ?, ?, ?, ?)",
        (staff_id, shift_date, start_time, end_time, memo)
    )

    d = date.fromisoformat(shift_date)
    return redirect(url_for("admin", year=d.year, month=d.month))


# ── シフト削除 ────────────────────────────────────────────────────────

@app.route("/shift/delete/<int:shift_id>", methods=["POST"])
def shift_delete(shift_id):
    r = require_login()
    if r:
        return r

    row = query("SELECT shift_date FROM shift_records WHERE id=?", (shift_id,))
    execute("DELETE FROM shift_records WHERE id=?", (shift_id,))

    if row:
        d = date.fromisoformat(row[0]["shift_date"])
        return redirect(url_for("admin", year=d.year, month=d.month))
    return redirect(url_for("admin"))


# ── シフト編集 ────────────────────────────────────────────────────────

@app.route("/shift/edit/<int:shift_id>", methods=["GET", "POST"])
def shift_edit(shift_id):
    r = require_login()
    if r:
        return r

    if request.method == "POST":
        start_time = request.form.get("start_time", "").strip()
        end_time   = request.form.get("end_time", "").strip()
        memo       = request.form.get("memo", "").strip()
        execute(
            "UPDATE shift_records SET start_time=?, end_time=?, memo=? WHERE id=?",
            (start_time, end_time, memo, shift_id)
        )
        row = query("SELECT shift_date FROM shift_records WHERE id=?", (shift_id,))
        if row:
            d = date.fromisoformat(row[0]["shift_date"])
            return redirect(url_for("admin", year=d.year, month=d.month))
        return redirect(url_for("admin"))

    shift = query(
        "SELECT s.*, st.name as staff_name FROM shift_records s "
        "JOIN shift_staff st ON s.staff_id = st.id WHERE s.id=?",
        (shift_id,)
    )
    if not shift:
        return redirect(url_for("admin"))
    staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
    return render_template("edit_shift.html", shift=shift[0], staff_rows=staff_rows)


# ── スタッフ管理 ──────────────────────────────────────────────────────

COLORS = [
    "#c87941", "#4a90d9", "#6ab04c", "#e056a0",
    "#f39c12", "#8e44ad", "#16a085", "#c0392b",
    "#2980b9", "#27ae60", "#d35400", "#7f8c8d",
]

@app.route("/staff", methods=["GET", "POST"])
def staff():
    r = require_login()
    if r:
        return r

    msg = error = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            name = request.form.get("name", "").strip()
            if not name:
                error = "名前を入力してください"
            else:
                existing = query("SELECT * FROM shift_staff")
                color = COLORS[len(existing) % len(COLORS)]
                allowed_slots   = request.form.get("allowed_slots", "").strip()
                max_days        = int(request.form.get("max_days", 0) or 0)
                wed_ok          = 1 if request.form.get("wed_ok") else 0
                sun_ok          = 1 if request.form.get("sun_ok") else 0
                sun_a_exclusive = 1 if request.form.get("sun_a_exclusive") else 0
                same_day_ng     = request.form.get("same_day_ng", "").strip()
                days_off_str    = request.form.get("days_off_str", "").strip()
                try:
                    execute(
                        "INSERT INTO shift_staff(name, color, allowed_slots, max_days, wed_ok, sun_ok, "
                        "sun_a_exclusive, same_day_ng, days_off_str) VALUES (?,?,?,?,?,?,?,?,?)",
                        (name, color, allowed_slots, max_days, wed_ok, sun_ok,
                         sun_a_exclusive, same_day_ng, days_off_str)
                    )
                    msg = f"{name} を追加しました"
                except Exception:
                    error = "その名前は既に登録されています"

        elif action == "update":
            sid             = request.form.get("staff_id")
            allowed_slots   = request.form.get("allowed_slots", "").strip()
            max_days        = int(request.form.get("max_days", 0) or 0)
            min_sun_days    = int(request.form.get("min_sun_days", 0) or 0)
            wed_ok          = 1 if request.form.get("wed_ok") else 0
            sun_ok          = 1 if request.form.get("sun_ok") else 0
            sun_only        = 1 if request.form.get("sun_only") else 0
            sun_a_exclusive = 1 if request.form.get("sun_a_exclusive") else 0
            wants_more      = 1 if request.form.get("wants_more") else 0
            same_day_ng     = request.form.get("same_day_ng", "").strip()
            days_off_str    = request.form.get("days_off_str", "").strip()
            color           = request.form.get("color", "#c87941")
            execute(
                "UPDATE shift_staff SET color=?, allowed_slots=?, max_days=?, min_sun_days=?, wed_ok=?, sun_ok=?, "
                "sun_only=?, sun_a_exclusive=?, wants_more=?, same_day_ng=?, days_off_str=? WHERE id=?",
                (color, allowed_slots, max_days, min_sun_days, wed_ok, sun_ok,
                 sun_only, sun_a_exclusive, wants_more, same_day_ng, days_off_str, sid)
            )
            msg = "スタッフ情報を更新しました"

        elif action == "delete":
            sid = request.form.get("staff_id")
            name_row = query("SELECT name FROM shift_staff WHERE id=?", (sid,))
            if name_row:
                name = name_row[0]["name"]
                execute("DELETE FROM shift_records WHERE staff_id=?", (sid,))
                execute("DELETE FROM shift_requests WHERE staff_name=?", (name,))
                execute("DELETE FROM shift_staff WHERE id=?", (sid,))
                msg = f"{name} を削除しました"

    staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
    return render_template("staff.html", staff_rows=staff_rows, msg=msg, error=error, colors=COLORS)


# ── 曜日別シフト設定 ─────────────────────────────────────────────────

DOW_JP = ["月", "火", "水", "木", "金", "土", "日"]

@app.route("/day-rules", methods=["GET", "POST"])
def day_rules():
    r = require_login()
    if r:
        return r

    msg = None
    if request.method == "POST":
        for dow in range(7):
            is_closed  = 1 if request.form.get(f"closed_{dow}") else 0
            slot_a     = int(request.form.get(f"a_{dow}", 0) or 0)
            slot_b     = int(request.form.get(f"b_{dow}", 0) or 0)
            slot_c     = int(request.form.get(f"c_{dow}", 0) or 0)
            slot_d     = int(request.form.get(f"d_{dow}", 0) or 0)
            slot_m     = int(request.form.get(f"m_{dow}", 0) or 0)
            slot_shikomi = int(request.form.get(f"shikomi_{dow}", 0) or 0)
            existing = query("SELECT dow FROM shift_day_rules WHERE dow=?", (dow,))
            if existing:
                execute(
                    "UPDATE shift_day_rules SET is_closed=?,slot_a=?,slot_b=?,slot_c=?,slot_d=?,slot_m=?,slot_shikomi=? WHERE dow=?",
                    (is_closed, slot_a, slot_b, slot_c, slot_d, slot_m, slot_shikomi, dow)
                )
            else:
                execute(
                    "INSERT INTO shift_day_rules (dow,is_closed,slot_a,slot_b,slot_c,slot_d,slot_m,slot_shikomi) VALUES (?,?,?,?,?,?,?,?)",
                    (dow, is_closed, slot_a, slot_b, slot_c, slot_d, slot_m, slot_shikomi)
                )
        msg = "曜日設定を保存しました"

    rows = {r["dow"]: r for r in query("SELECT * FROM shift_day_rules ORDER BY dow")}
    return render_template("day_rules.html", rows=rows, dow_jp=DOW_JP, msg=msg)


# ── 自動生成 ──────────────────────────────────────────────────────────

@app.route("/generate", methods=["GET", "POST"])
def generate():
    r = require_login()
    if r:
        return r

    today = today_jst()
    msg = error = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "generate":
            year  = int(request.form.get("year",  today.year))
            month = int(request.form.get("month", today.month))
            marche_dates_str = request.form.get("marche_dates", "").strip()
            holiday_dates_str = request.form.get("holiday_dates", "").strip()
            holiday_dates = set()
            for x in holiday_dates_str.split(","):
                x = x.strip()
                if x.isdigit():
                    holiday_dates.add(int(x))

            # 設定保存
            month_key = f"{year}-{month:02d}"
            existing_setting = query("SELECT id FROM shift_config WHERE month=?", (month_key,))
            if existing_setting:
                execute(
                    "UPDATE shift_config SET marche_dates=? WHERE month=?",
                    (marche_dates_str, month_key)
                )
            else:
                execute(
                    "INSERT INTO shift_config (month, marche_dates) VALUES (?, ?)",
                    (month_key, marche_dates_str)
                )

            marche_dates = set()
            for x in marche_dates_str.split(","):
                x = x.strip()
                if x.isdigit():
                    marche_dates.add(int(x))

            staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
            if not staff_rows:
                error = "スタッフが登録されていません"
            else:
                # シフト希望をdays_off_strに反映
                requests = query(
                    "SELECT staff_name, days_off_str FROM shift_requests WHERE month=?",
                    (month_key,)
                )
                req_map = {r["staff_name"]: r["days_off_str"] for r in requests}
                for s in staff_rows:
                    if s["name"] in req_map:
                        s["days_off_str"] = req_map[s["name"]]
                # 既存の自動生成シフトを削除（月全体）
                first_day = date(year, month, 1)
                last_day  = date(year, month, calendar.monthrange(year, month)[1])
                execute(
                    "DELETE FROM shift_records WHERE shift_date BETWEEN ? AND ?",
                    (first_day.isoformat(), last_day.isoformat())
                )

                assignments = _generate_shifts(year, month, staff_rows, marche_dates, holiday_dates)

                # staff名→IDマップ
                sid_map = {s["name"]: s["id"] for s in staff_rows}

                for (date_str, slot, name, start_t, end_t, memo) in assignments:
                    sid = sid_map.get(name)
                    if sid:
                        execute(
                            "INSERT INTO shift_records (staff_id, shift_date, start_time, end_time, slot_label, memo) "
                            "VALUES (?,?,?,?,?,?)",
                            (sid, date_str, start_t, end_t, slot, memo)
                        )

                msg = f"{year}年{month}月のシフトを生成しました（{len(assignments)}件）"
                return redirect(url_for("admin", year=year, month=month))

    staff_rows = query("SELECT * FROM shift_staff ORDER BY id")
    # 直近3ヶ月の提出済み希望を表示
    requests = query(
        "SELECT * FROM shift_requests ORDER BY month DESC, staff_name"
    )
    return render_template("generate.html",
        staff_rows=staff_rows, msg=msg, error=error,
        today=today, requests=requests,
    )


def _generate_shifts(year, month, staff_rows, marche_dates, holiday_dates=None):
    """
    シフト自動生成アルゴリズム
    Returns: list of (date_str, slot, name, start_time, end_time, memo)
    """
    SLOT_TIMES = {
        "A":     ("06:30", "12:30"),
        "B":     ("09:00", "15:00"),
        "C":     ("09:00", "15:00"),
        "D":     ("10:00", "16:00"),
        "M":     ("07:00", "13:00"),
        "仕込み": ("09:00", "15:00"),
    }

    # 曜日ごとの設定を取得
    day_rules_rows = query("SELECT * FROM shift_day_rules ORDER BY dow")
    day_rules = {r["dow"]: r for r in day_rules_rows}

    # スタッフ情報を整理
    staff_info = []
    for s in staff_rows:
        allowed = [x.strip() for x in (s.get("allowed_slots") or "").split(",") if x.strip()]
        days_off = set()
        for x in (s.get("days_off_str") or "").split(","):
            x = x.strip()
            if x.isdigit():
                days_off.add(int(x))
        same_day_ng = [x.strip() for x in (s.get("same_day_ng") or "").split(",") if x.strip()]
        staff_info.append({
            "id":              s["id"],
            "name":            s["name"],
            "allowed_slots":   allowed,
            "max_days":        s.get("max_days") or 0,
            "min_sun_days":    s.get("min_sun_days") or 0,
            "wed_ok":          bool(s.get("wed_ok")),
            "sun_ok":          bool(s.get("sun_ok")),
            "sun_a_exclusive": bool(s.get("sun_a_exclusive")),
            "sun_only":        bool(s.get("sun_only")),
            "wants_more":      bool(s.get("wants_more")),
            "same_day_ng":     same_day_ng,
            "days_off":        days_off,
        })

    days_in_month = calendar.monthrange(year, month)[1]
    work_count     = {s["name"]: 0 for s in staff_info}
    sun_work_count = {s["name"]: 0 for s in staff_info}
    assignments    = []

    for day in range(1, days_in_month + 1):
        d   = date(year, month, day)
        dow = d.weekday()  # 0=月...6=日

        # 休業日（管理者指定）はシフトを作らない
        if holiday_dates and day in holiday_dates:
            continue

        rule = day_rules.get(dow, {})
        if rule.get("is_closed", dow == 5):
            continue

        # 曜日設定から必要枠リストを生成
        slots_needed = []
        slots_needed += ["A"]    * int(rule.get("slot_a", 0) or 0)
        slots_needed += ["B"]    * int(rule.get("slot_b", 0) or 0)
        slots_needed += ["C"]    * int(rule.get("slot_c", 0) or 0)
        slots_needed += ["D"]    * int(rule.get("slot_d", 0) or 0)
        slots_needed += ["仕込み"] * int(rule.get("slot_shikomi", 0) or 0)
        # 日曜のマルシェ日はM枠を追加
        if dow == 6 and day in marche_dates:
            slots_needed += ["M"] * max(int(rule.get("slot_m", 0) or 0), 2)
        else:
            slots_needed += ["M"] * int(rule.get("slot_m", 0) or 0)

        if not slots_needed:
            continue

        assigned_today = {}   # slot_key -> name
        assigned_names = set()

        for slot in slots_needed:
            slot_key = slot if slot != "M" else f"M{len([k for k in assigned_today if k.startswith('M')])}"

            eligible = []
            for s in staff_info:
                name = s["name"]
                if name in assigned_names:
                    continue
                if day in s["days_off"]:
                    continue
                if s["max_days"] > 0 and work_count[name] >= s["max_days"]:
                    continue
                # 日曜のみ出勤スタッフは日曜以外スキップ
                if s["sun_only"] and dow != 6:
                    continue

                # 枠適格チェック
                ok = False
                if slot == "仕込み":
                    # B or C枠の人で水曜可
                    ok = s["wed_ok"] and (
                        "B" in s["allowed_slots"] or "C" in s["allowed_slots"]
                    )
                elif slot == "A":
                    ok = "A" in s["allowed_slots"]
                elif slot == "M":
                    ok = "B" in s["allowed_slots"] or "C" in s["allowed_slots"] or "A" in s["allowed_slots"]
                elif slot in ("B", "C"):
                    ok = slot in s["allowed_slots"]
                elif slot == "D":
                    ok = "D" in s["allowed_slots"]

                if not ok:
                    continue

                eligible.append(s)

            if not eligible:
                continue

            # 優先度ソート
            def priority(s):
                cnt = work_count[s["name"]]
                # 日曜：月1回以上 or min_sun_days 未達成のスタッフを最優先
                sun_needed = 1
                if dow == 6:
                    min_req = max(s["min_sun_days"], 1)  # 最低1回は日曜に入れる
                    sun_needed = 0 if sun_work_count[s["name"]] < min_req else 1
                # 日曜A枠は sun_a_exclusive を優先
                exclusive_bonus = 0 if (slot == "A" and dow == 6 and s["sun_a_exclusive"]) else 1
                # 週3以上希望のスタッフを常に優先（12日未達なら最優先、達成後も優先）
                if s["wants_more"]:
                    more_needed = 0 if cnt < 12 else 1
                else:
                    more_needed = 2
                return (more_needed, sun_needed, exclusive_bonus, cnt)

            eligible.sort(key=priority)
            chosen = eligible[0]["name"]

            assigned_today[slot_key] = chosen
            assigned_names.add(chosen)
            work_count[chosen] += 1
            if dow == 6:
                sun_work_count[chosen] += 1

            start_t, end_t = SLOT_TIMES.get(slot, ("09:00", "17:00"))
            memo = f"{slot}枠" if slot != "仕込み" else "仕込み"
            assignments.append((d.isoformat(), slot, chosen, start_t, end_t, memo))

    return assignments


# ── Excel出力 ────────────────────────────────────────────────────────

@app.route("/export/excel/<int:year>/<int:month>")
def export_excel(year, month):
    r = require_login()
    if r:
        return r

    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
    except ImportError:
        return "openpyxl がインストールされていません", 500

    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])

    shifts_raw = query(
        "SELECT s.*, st.name as staff_name FROM shift_records s "
        "JOIN shift_staff st ON s.staff_id = st.id "
        "WHERE s.shift_date BETWEEN ? AND ? ORDER BY s.shift_date, s.slot_label",
        (first_day.isoformat(), last_day.isoformat())
    )

    # 日付 → {slot: name} マップ
    day_map = {}
    for sh in shifts_raw:
        d = sh["shift_date"]
        slot = sh.get("slot_label") or sh.get("memo") or ""
        name = sh["staff_name"]
        day_map.setdefault(d, {})[slot] = name

    setting = query("SELECT * FROM shift_config WHERE month=?", (f"{year}-{month:02d}",))
    marche_dates = set()
    if setting:
        for x in (setting[0].get("marche_dates") or "").split(","):
            x = x.strip()
            if x.isdigit():
                marche_dates.add(int(x))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シフト"

    # ── スタイル定義 ──
    dark_fill   = PatternFill("solid", fgColor="2D2D2D")
    sun_fill    = PatternFill("solid", fgColor="FFE0E0")
    marche_fill = PatternFill("solid", fgColor="E8F0E8")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    body_font   = Font(size=9)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="BBBBBB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    DOW_JP = ["月", "火", "水", "木", "金", "土", "日"]

    # ── ヘッダー行 ──
    headers = [
        ("日付", 8), ("曜日", 5),
        ("A枠\n6:30〜\n製造・品出し", 12),
        ("B枠\n9:00〜\nレジ", 12),
        ("C枠\n9:00〜\nレジ補助", 12),
        ("D枠\n10:00〜\n品出し(日曜)", 12),
        ("M枠\n7:00〜\nマルシェ", 14),
        ("備考", 20),
    ]
    for col_idx, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = dark_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[chr(64 + col_idx)].width = w

    ws.row_dimensions[1].height = 45

    # ── データ行 ──
    row_num = 2
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        d   = date(year, month, day)
        dow = d.weekday()
        if dow == 5:  # 土曜スキップ
            continue

        date_str = d.isoformat()
        slots    = day_map.get(date_str, {})

        is_sun    = dow == 6
        is_marche = is_sun and day in marche_dates

        # 備考（仕込みの場合）
        notes = []
        if dow in (1, 2):
            name = slots.get("仕込み", "")
            if name:
                notes.append(f"{name} (9:00〜仕込み)")

        # 休み希望の備考
        staff_rows = query("SELECT name, days_off_str FROM shift_staff ORDER BY id")
        off_names = []
        for s in staff_rows:
            offs = set()
            for x in (s.get("days_off_str") or "").split(","):
                x = x.strip()
                if x.isdigit():
                    offs.add(int(x))
            if day in offs:
                off_names.append(s["name"])
        if off_names:
            notes.append("・".join(off_names) + " 休み希望")
        if is_marche:
            notes.append("🛒 マルシェ")

        values = [
            f"{month}/{day}",
            DOW_JP[dow],
            slots.get("A", "") if dow not in (1, 2) else "",
            slots.get("B", "") if dow not in (1, 2) else ("\n".join(notes) if dow in (1, 2) else ""),
            slots.get("C", "") if dow not in (1, 2) else "",
            slots.get("D", "") if is_sun else "",
            (slots.get("M0", "") + "\n" + slots.get("M1", "")).strip() if is_marche else "",
            ("\n".join(notes)) if dow not in (1, 2) else "",
        ]

        # 仕込み日は B列にまとめて表示
        if dow in (1, 2):
            values[2] = ""
            values[3] = "\n".join(notes) if notes else ""
            values[4] = ""
            values[5] = ""

        fill = sun_fill if is_sun else (marche_fill if is_marche else None)

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font      = body_font
            cell.alignment = center
            cell.border    = border
            if fill:
                cell.fill = fill

        # 仕込み日はB〜E結合
        if dow in (1, 2) and any(notes):
            ws.merge_cells(
                start_row=row_num, start_column=3,
                end_row=row_num,   end_column=6
            )
            merged = ws.cell(row=row_num, column=3)
            merged.value     = "\n".join(notes)
            merged.font      = body_font
            merged.alignment = center
            merged.border    = border

        ws.row_dimensions[row_num].height = 22
        row_num += 1

    # 出力
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"シフト表_{year}年{month}月.xlsx"
    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── 日別JSON ────────────────────────────────────────────────────────

@app.route("/api/day")
def api_day():
    r = require_login()
    if r:
        return jsonify([])
    d = request.args.get("date", "")
    rows = query(
        "SELECT s.*, st.name as staff_name, st.color FROM shift_records s "
        "JOIN shift_staff st ON s.staff_id = st.id "
        "WHERE s.shift_date=? ORDER BY s.start_time",
        (d,)
    )
    return jsonify(rows)


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5050, host="0.0.0.0")

init_db()
