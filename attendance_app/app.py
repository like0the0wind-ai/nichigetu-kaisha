import io
import os
from collections import defaultdict
from datetime import datetime, date, timezone, timedelta
from dateutil.relativedelta import relativedelta
from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify

JST = timezone(timedelta(hours=9))

def now_jst():
    return datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")

def today_jst():
    return datetime.now(JST).strftime("%Y-%m-%d")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "nichigetsu-secret")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "saito")

DATABASE_URL = os.environ.get("DATABASE_URL")

# ── DB ────────────────────────────────────────────────────────────

def get_db():
    if DATABASE_URL:
        import psycopg2, psycopg2.extras
        url = DATABASE_URL.replace("postgres://", "postgresql://", 1)
        conn = psycopg2.connect(url)
        conn.cursor_factory = psycopg2.extras.RealDictCursor
        return conn
    else:
        import sqlite3
        db_path = os.path.join(os.path.dirname(__file__), "attendance.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn

def ph():
    return "%s" if DATABASE_URL else "?"

def init_db():
    conn = get_db()
    try:
        cur = conn.cursor()
        serial = "SERIAL" if DATABASE_URL else "INTEGER"
        ai     = "" if DATABASE_URL else "AUTOINCREMENT"

        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS records (
                id        {serial} PRIMARY KEY {ai},
                name      TEXT NOT NULL,
                clock_in  TEXT NOT NULL,
                clock_out TEXT,
                break_min INTEGER DEFAULT 0
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS breaks (
                id          {serial} PRIMARY KEY {ai},
                record_id   INTEGER NOT NULL,
                break_start TEXT NOT NULL,
                break_end   TEXT
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS staff (
                id   {serial} PRIMARY KEY {ai},
                name TEXT NOT NULL UNIQUE
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS sales (
                id       {serial} PRIMARY KEY {ai},
                date     TEXT NOT NULL,
                amount   INTEGER NOT NULL,
                category TEXT DEFAULT '',
                memo     TEXT DEFAULT ''
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS shifts (
                id         {serial} PRIMARY KEY {ai},
                name       TEXT NOT NULL,
                date       TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time   TEXT NOT NULL
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS products (
                id       {serial} PRIMARY KEY {ai},
                name     TEXT NOT NULL UNIQUE,
                category TEXT DEFAULT '',
                price    INTEGER DEFAULT 0,
                cost     INTEGER DEFAULT 0,
                recipe   TEXT DEFAULT ''
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS employees (
                id                  {serial} PRIMARY KEY {ai},
                name                TEXT NOT NULL UNIQUE,
                hourly_rate         INTEGER NOT NULL DEFAULT 0,
                transport_allowance INTEGER NOT NULL DEFAULT 0,
                other_allowance     INTEGER NOT NULL DEFAULT 0,
                notes               TEXT,
                fuyou_target        INTEGER NOT NULL DEFAULT 0
            )
        """)
        for col_sql in [
            "ALTER TABLE employees ADD COLUMN fuyou_target INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE employees ADD COLUMN other_income INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE employees ADD COLUMN fuyou_category TEXT NOT NULL DEFAULT '103'",
        ]:
            try:
                cur.execute(col_sql)
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS employee_history (
                id                  {serial} PRIMARY KEY {ai},
                name                TEXT NOT NULL,
                hourly_rate         INTEGER NOT NULL DEFAULT 0,
                transport_allowance INTEGER NOT NULL DEFAULT 0,
                other_allowance     INTEGER NOT NULL DEFAULT 0,
                other_income        INTEGER NOT NULL DEFAULT 0,
                fuyou_category      TEXT NOT NULL DEFAULT '103',
                effective_from      TEXT NOT NULL
            )
        """)
        conn.commit()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS payslips (
                id                   {serial} PRIMARY KEY {ai},
                employee_name        TEXT NOT NULL,
                period_start         TEXT NOT NULL,
                period_end           TEXT NOT NULL,
                total_work_min       INTEGER NOT NULL DEFAULT 0,
                regular_min          INTEGER NOT NULL DEFAULT 0,
                overtime_min         INTEGER NOT NULL DEFAULT 0,
                base_pay             INTEGER NOT NULL DEFAULT 0,
                overtime_pay         INTEGER NOT NULL DEFAULT 0,
                transport_allowance  INTEGER NOT NULL DEFAULT 0,
                other_allowance      INTEGER NOT NULL DEFAULT 0,
                gross_pay            INTEGER NOT NULL DEFAULT 0,
                health_insurance     INTEGER NOT NULL DEFAULT 0,
                pension              INTEGER NOT NULL DEFAULT 0,
                employment_insurance INTEGER NOT NULL DEFAULT 0,
                income_tax           INTEGER NOT NULL DEFAULT 0,
                other_deduction      INTEGER NOT NULL DEFAULT 0,
                total_deduction      INTEGER NOT NULL DEFAULT 0,
                net_pay              INTEGER NOT NULL DEFAULT 0,
                note                 TEXT,
                created_at           TEXT NOT NULL
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS year_end_adj (
                id                   {serial} PRIMARY KEY {ai},
                name                 TEXT NOT NULL,
                year                 INTEGER NOT NULL,
                social_insurance     INTEGER NOT NULL DEFAULT 0,
                life_insurance       INTEGER NOT NULL DEFAULT 0,
                earthquake_insurance INTEGER NOT NULL DEFAULT 0,
                disability           INTEGER NOT NULL DEFAULT 0,
                dependent_count      INTEGER NOT NULL DEFAULT 0,
                specific_dependent   INTEGER NOT NULL DEFAULT 0,
                old_dependent        INTEGER NOT NULL DEFAULT 0,
                spouse_income        INTEGER NOT NULL DEFAULT 0,
                withheld_tax         INTEGER NOT NULL DEFAULT 0,
                UNIQUE(name, year)
            )
        """)
        conn.commit()
    finally:
        conn.close()

def query(sql, params=()):
    p = ph()
    sql = sql.replace("?", p)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        conn.commit()
        try:
            return cur.fetchall()
        except Exception:
            return []
    finally:
        conn.close()

def execute(sql, params=()):
    p = ph()
    sql = sql.replace("?", p)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        conn.commit()
    finally:
        conn.close()

FUYOU_LIMITS = {"103": 1_030_000, "123": 1_230_000}

def get_emp_at(name, on_date):
    """指定日に有効だった従業員の時給・扶養区分などを返す（履歴優先、なければ現在の設定）"""
    rows = query(
        "SELECT * FROM employee_history WHERE name=? AND effective_from<=? ORDER BY effective_from DESC, id DESC LIMIT 1",
        (name, on_date)
    )
    if rows:
        return dict(rows[0])
    cur = query("SELECT * FROM employees WHERE name=?", (name,))
    if cur:
        e = dict(cur[0])
        e.setdefault("fuyou_category", "103")
        e.setdefault("other_income", 0)
        return e
    return None

def record_employee_history(name, hourly, transport, other, other_income, fuyou_category):
    """設定が変わった場合のみ履歴に新しいスナップショットを追加する"""
    today = today_jst()
    existing = query(
        "SELECT * FROM employee_history WHERE name=? ORDER BY effective_from DESC, id DESC LIMIT 1",
        (name,)
    )
    if existing:
        e = existing[0]
        if (int(e["hourly_rate"]) == hourly and int(e["transport_allowance"]) == transport and
            int(e["other_allowance"]) == other and int(e["other_income"] or 0) == other_income and
            (e["fuyou_category"] or "103") == fuyou_category):
            return
    execute(
        "INSERT INTO employee_history (name, hourly_rate, transport_allowance, other_allowance, other_income, fuyou_category, effective_from) VALUES (?,?,?,?,?,?,?)",
        (name, hourly, transport, other, other_income, fuyou_category, today)
    )

STANDARD_HOURS = 6

def ceil15(minutes):
    import math
    return math.ceil(minutes / 15) * 15

def floor15(minutes):
    return (minutes // 15) * 15

def fmt_time(minutes):
    h = minutes // 60
    m = minutes % 60
    if m == 0:
        return f"{h}h"
    return f"{h}h{m:02d}"

# ── 従業員側（打刻） ──────────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def index():
    msg = None
    if request.method == "POST":
        action = request.form["action"]
        name   = request.form["name"].strip()
        if not name:
            msg = "名前を入力してください"
        elif action == "clock_in":
            rows = query("SELECT id FROM records WHERE name=? AND clock_out IS NULL", (name,))
            if rows:
                msg = f"{name} さんはすでに出勤中です"
            else:
                execute("INSERT INTO records (name, clock_in) VALUES (?, ?)", (name, now_jst()))
                msg = f"{name} さんの出勤を記録しました ✓"
        elif action == "clock_out":
            rows = query("SELECT id FROM records WHERE name=? AND clock_out IS NULL", (name,))
            if not rows:
                msg = f"{name} さんの出勤記録が見つかりません"
            else:
                rec_id = rows[0]["id"]
                br = query("SELECT id FROM breaks WHERE record_id=? AND break_end IS NULL", (rec_id,))
                if br:
                    execute("UPDATE breaks SET break_end=? WHERE id=?", (now_jst(), br[0]["id"]))
                brs = query("SELECT break_start, break_end FROM breaks WHERE record_id=?", (rec_id,))
                total_break = 0
                for b in brs:
                    if b["break_end"]:
                        s = datetime.strptime(b["break_start"], "%Y-%m-%d %H:%M:%S")
                        e = datetime.strptime(b["break_end"],   "%Y-%m-%d %H:%M:%S")
                        total_break += int((e - s).total_seconds() // 60)
                execute("UPDATE records SET clock_out=?, break_min=? WHERE id=?",
                        (now_jst(), total_break, rec_id))
                msg = f"{name} さんの退勤を記録しました ✓"
        elif action == "break_start":
            rows = query("SELECT id FROM records WHERE name=? AND clock_out IS NULL", (name,))
            if not rows:
                msg = f"{name} さんの出勤記録が見つかりません"
            else:
                rec_id = rows[0]["id"]
                br = query("SELECT id FROM breaks WHERE record_id=? AND break_end IS NULL", (rec_id,))
                if br:
                    msg = f"{name} さんはすでに休憩中です"
                else:
                    execute("INSERT INTO breaks (record_id, break_start) VALUES (?, ?)",
                            (rec_id, now_jst()))
                    msg = f"{name} さんの休憩開始を記録しました ✓"
        elif action == "break_end":
            rows = query("SELECT id FROM records WHERE name=? AND clock_out IS NULL", (name,))
            if not rows:
                msg = f"{name} さんの出勤記録が見つかりません"
            else:
                rec_id = rows[0]["id"]
                br = query("SELECT id FROM breaks WHERE record_id=? AND break_end IS NULL", (rec_id,))
                if not br:
                    msg = f"{name} さんは休憩中ではありません"
                else:
                    execute("UPDATE breaks SET break_end=? WHERE id=?", (now_jst(), br[0]["id"]))
                    msg = f"{name} さんの休憩終了を記録しました ✓"

    staff_rows = query("SELECT name FROM staff ORDER BY name")
    staff_status = []
    for s in staff_rows:
        name = s["name"]
        rec = query("SELECT id FROM records WHERE name=? AND clock_out IS NULL", (name,))
        if rec:
            br = query("SELECT id FROM breaks WHERE record_id=? AND break_end IS NULL", (rec[0]["id"],))
            status = "break" if br else "in"
        else:
            status = "out"
        staff_status.append({"name": name, "status": status})
    return render_template("index.html", staff_status=staff_status, msg=msg)

# ── 給与期間 ──────────────────────────────────────────────────────

def current_pay_period(today=None):
    t = today or date.today()
    if t.day >= 11:
        return t.replace(day=11), (t + relativedelta(months=1)).replace(day=10)
    return (t - relativedelta(months=1)).replace(day=11), t.replace(day=10)

def pay_period_label(start, end):
    return f"{start.month}月{start.day}日〜{end.month}月{end.day}日"

# ── 管理者側 ──────────────────────────────────────────────────────

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

@app.route("/admin")
@admin_required
def admin():
    default_start, default_end = current_pay_period()
    date_from = request.args.get("from", default_start.isoformat())
    date_to   = request.args.get("to",   default_end.isoformat())

    rows = query(
        "SELECT * FROM records WHERE date(clock_in) BETWEEN ? AND ? ORDER BY clock_in",
        (date_from, date_to)
    )

    staff, records = {}, []
    for r in rows:
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S") if r["clock_out"] else None
        break_min = r["break_min"] or 0

        if co:
            ci_min = ceil15(ci.hour * 60 + ci.minute)
            co_min = floor15(co.hour * 60 + co.minute)
            total_min = max(0, co_min - ci_min - break_min)
            std_min  = min(total_min, STANDARD_HOURS * 60)
            over_min = floor15(max(0, total_min - STANDARD_HOURS * 60))
        else:
            std_min = over_min = None
            total_min = 0

        records.append({
            "id": r["id"], "name": r["name"],
            "date": ci.strftime("%m/%d"),
            "clock_in": ci.strftime("%H:%M"),
            "clock_out": co.strftime("%H:%M") if co else "—",
            "work": fmt_time(std_min) if std_min is not None else "出勤中",
            "overtime": fmt_time(over_min) if over_min else "—",
            "break": fmt_time(break_min) if break_min else "—",
            "work_min": std_min or 0,
            "over_min": over_min or 0,
        })
        if r["name"] not in staff:
            staff[r["name"]] = {"days": 0, "total_min": 0, "over_min": 0, "break_min": 0}
        staff[r["name"]]["days"] += 1
        staff[r["name"]]["total_min"] += std_min or 0
        staff[r["name"]]["over_min"] += over_min or 0
        staff[r["name"]]["break_min"] += break_min

    # 年間収入計算（扶養チェック用）
    year_start = date.fromisoformat(date_from).replace(month=1, day=1).isoformat()
    year_end   = date.fromisoformat(date_to).replace(month=12, day=31).isoformat()

    staff_summary = []
    for n, s in sorted(staff.items()):
        emp_rows = query("SELECT * FROM employees WHERE name=?", (n,))
        emp = dict(emp_rows[0]) if emp_rows else None
        if emp:
            emp.setdefault("fuyou_target", 0)
            emp.setdefault("other_income", 0)
            emp.setdefault("fuyou_category", "103")
        category = (emp["fuyou_category"] if emp and emp.get("fuyou_category") else "103")
        fuyou_limit = FUYOU_LIMITS.get(category) if category != "none" else None

        # 年間レコード取得（記録ごとに当時の時給・設定を使って計算）
        yr_rows = query(
            "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? AND clock_out IS NOT NULL",
            (n, year_start, year_end)
        )
        year_total = 0
        for rr in yr_rows:
            ci = datetime.strptime(rr["clock_in"], "%Y-%m-%d %H:%M:%S")
            co = datetime.strptime(rr["clock_out"], "%Y-%m-%d %H:%M:%S")
            bm = rr["break_min"] or 0
            ci_m = ceil15(ci.hour * 60 + ci.minute)
            co_m = floor15(co.hour * 60 + co.minute)
            tot  = max(0, co_m - ci_m - bm)
            ov   = max(0, tot - STANDARD_HOURS * 60)
            reg  = tot - ov

            rec_date = ci.strftime("%Y-%m-%d")
            emp_at = get_emp_at(n, rec_date)
            hourly_at    = int(emp_at["hourly_rate"] or 0) if emp_at else 0
            transport_at = int(emp_at["transport_allowance"] or 0) if emp_at else 0

            pay = int(hourly_at * reg / 60) + int(hourly_at * ov / 60) + transport_at
            year_total += pay

        other_income = int(emp["other_income"] or 0) if emp else 0
        year_total += other_income
        if fuyou_limit:
            remaining = fuyou_limit - year_total
            pct = min(100, int(year_total / fuyou_limit * 100))
        else:
            remaining = None
            pct = 0

        staff_summary.append({
            "name": n, "days": s["days"],
            "total": fmt_time(s["total_min"]),
            "overtime": fmt_time(s["over_min"]) if s["over_min"] else "—",
            "total_min": s["total_min"], "over_min": s["over_min"],
            "year_total": year_total,
            "remaining": remaining,
            "pct": pct,
            "fuyou_category": category,
            "fuyou_limit": fuyou_limit,
            "other_income": other_income,
        })

    return render_template("admin.html",
        staff_summary=staff_summary,
        date_from=date_from, date_to=date_to,
        period_label=pay_period_label(
            date.fromisoformat(date_from), date.fromisoformat(date_to)))

@app.route("/admin/attendance")
@admin_required
def admin_attendance():
    default_start, default_end = current_pay_period()
    date_from = request.args.get("from", default_start.isoformat())
    date_to   = request.args.get("to",   default_end.isoformat())

    rows = query(
        "SELECT * FROM records WHERE date(clock_in) BETWEEN ? AND ? ORDER BY clock_in",
        (date_from, date_to)
    )
    records = []
    for r in rows:
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S") if r["clock_out"] else None
        break_min = r["break_min"] or 0
        if co:
            ci_min = ceil15(ci.hour * 60 + ci.minute)
            co_min = floor15(co.hour * 60 + co.minute)
            total_min = max(0, co_min - ci_min - break_min)
            std_min  = min(total_min, STANDARD_HOURS * 60)
            over_min = floor15(max(0, total_min - STANDARD_HOURS * 60))
        else:
            std_min = over_min = None
        records.append({
            "id": r["id"], "name": r["name"],
            "date": ci.strftime("%m/%d"),
            "clock_in": ci.strftime("%H:%M"),
            "clock_out": co.strftime("%H:%M") if co else "—",
            "work": fmt_time(std_min) if std_min is not None else "出勤中",
            "overtime": fmt_time(over_min) if over_min else "—",
            "break": fmt_time(break_min) if break_min else "—",
            "work_min": std_min or 0,
            "over_min": over_min or 0,
        })

    staff_dict = defaultdict(list)
    for r in records:
        staff_dict[r["name"]].append(r)
    staff_records = [{"name": n, "records": staff_dict[n]} for n in sorted(staff_dict)]

    return render_template("attendance.html",
        staff_records=staff_records,
        date_from=date_from, date_to=date_to,
        period_label=pay_period_label(
            date.fromisoformat(date_from), date.fromisoformat(date_to)))

# ── スタッフ管理 ──────────────────────────────────────────────────

@app.route("/admin/staff", methods=["GET", "POST"])
@admin_required
def admin_staff():
    error = None
    if request.method == "POST":
        action = request.form.get("action")
        name   = request.form.get("name", "").strip()
        if action == "add" and name:
            try:
                execute("INSERT INTO staff (name) VALUES (?)", (name,))
            except Exception:
                error = f"「{name}」はすでに登録されています"
        elif action == "delete" and name:
            execute("DELETE FROM staff WHERE name=?", (name,))
    staff = query("SELECT * FROM staff ORDER BY name")
    return render_template("staff.html", staff=staff, error=error)

# ── 記録編集・削除 ────────────────────────────────────────────────

@app.route("/admin/add", methods=["GET", "POST"])
@admin_required
def admin_add():
    staff_rows = query("SELECT name FROM staff ORDER BY name")
    error = None
    if request.method == "POST":
        name      = request.form.get("name", "").strip()
        clock_in  = request.form.get("clock_in", "").strip()
        clock_out = request.form.get("clock_out", "").strip()
        break_min = int(request.form.get("break_min") or 0)
        back      = request.form.get("back", "")
        try:
            ci = datetime.strptime(clock_in, "%Y-%m-%dT%H:%M")
            co = datetime.strptime(clock_out, "%Y-%m-%dT%H:%M") if clock_out else None
            if co and co <= ci:
                raise ValueError("退勤が出勤以前の時刻です")
            execute(
                "INSERT INTO records (name, clock_in, clock_out, break_min) VALUES (?,?,?,?)",
                (name,
                 ci.strftime("%Y-%m-%d %H:%M:%S"),
                 co.strftime("%Y-%m-%d %H:%M:%S") if co else None,
                 break_min)
            )
            return redirect(back or url_for("admin_attendance"))
        except ValueError as e:
            error = f"入力エラー: {e}"
    back = request.args.get("back", "")
    prefill_name = request.args.get("name", "")
    prefill_date = request.args.get("date", today_jst())
    return render_template("add_record.html",
                           staff_rows=staff_rows, error=error,
                           back=back, prefill_name=prefill_name,
                           prefill_date=prefill_date)


@app.route("/admin/export/monthly")
@admin_required
def admin_export_monthly():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    year  = int(request.args.get("year",  datetime.now(JST).year))
    month = int(request.args.get("month", datetime.now(JST).month))
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    date_from = f"{year}-{month:02d}-01"
    date_to   = f"{year}-{month:02d}-{last_day}"

    rows = query(
        "SELECT * FROM records WHERE date(clock_in) BETWEEN ? AND ? ORDER BY name, clock_in",
        (date_from, date_to)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月"
    ws.page_setup.paperSize  = 9
    ws.page_setup.orientation = "portrait"

    hdr_fill = PatternFill("solid", fgColor="C8A878")
    hdr_font = Font(name="MS明朝", size=9, bold=True)
    val_font = Font(name="MS明朝", size=9)
    center   = Alignment(horizontal="center", vertical="center")
    right    = Alignment(horizontal="right",  vertical="center")

    headers = ["名前","日付","出勤","退勤","休憩(分)","勤務時間","残業"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    for col, w in enumerate([10,11,8,8,7,9,8], 1):
        ws.column_dimensions[chr(64+col)].width = w

    for i, r in enumerate(rows, 2):
        r = dict(r)
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S") if r["clock_out"] else None
        bm = r.get("break_min") or 0
        if co:
            ci_m = ceil15(ci.hour*60+ci.minute)
            co_m = floor15(co.hour*60+co.minute)
            total = max(0, co_m-ci_m-bm)
            std   = min(total, STANDARD_HOURS*60)
            over  = floor15(max(0, total-STANDARD_HOURS*60))
            work_str = fmt_time(std)
            over_str = fmt_time(over) if over else "—"
        else:
            work_str = "出勤中"
            over_str = "—"

        vals = [r["name"], ci.strftime("%Y/%m/%d"), ci.strftime("%H:%M"),
                co.strftime("%H:%M") if co else "—", bm, work_str, over_str]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = val_font
            cell.alignment = right if isinstance(v, int) else center

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"出退勤_{year}年{month}月.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/delete/<int:rec_id>", methods=["POST"])
@admin_required
def admin_delete(rec_id):
    execute("DELETE FROM records WHERE id=?", (rec_id,))
    back = request.form.get("back") or request.referrer or url_for("admin")
    return redirect(back)

@app.route("/admin/edit/<int:rec_id>", methods=["GET", "POST"])
@admin_required
def admin_edit(rec_id):
    rows = query("SELECT * FROM records WHERE id=?", (rec_id,))
    if not rows:
        return redirect(url_for("admin"))
    rec = rows[0]
    error = None
    if request.method == "POST":
        name      = request.form["name"].strip()
        clock_in  = request.form["clock_in"].strip()
        clock_out = request.form["clock_out"].strip()
        try:
            ci = datetime.strptime(clock_in, "%Y-%m-%dT%H:%M")
            co = datetime.strptime(clock_out, "%Y-%m-%dT%H:%M") if clock_out else None
            if co and co < ci:
                raise ValueError("退勤が出勤より早い")
            execute("UPDATE records SET name=?, clock_in=?, clock_out=? WHERE id=?",
                    (name,
                     ci.strftime("%Y-%m-%d %H:%M:%S"),
                     co.strftime("%Y-%m-%d %H:%M:%S") if co else None,
                     rec_id))
            back = request.form.get("back") or request.args.get("back") or url_for("admin")
            return redirect(back)
        except ValueError as e:
            error = f"入力エラー: {e}"

    ci_val = rec["clock_in"][:16].replace(" ", "T") if rec["clock_in"] else ""
    co_val = rec["clock_out"][:16].replace(" ", "T") if rec["clock_out"] else ""
    return render_template("edit.html", rec=rec, ci_val=ci_val, co_val=co_val, error=error)

# ── 年末調整 計算ロジック ──────────────────────────────────────────

def _emp_deduction(gross):
    """給与所得控除額（令和2年以降）"""
    if gross <= 1_625_000:
        return 550_000
    elif gross <= 1_800_000:
        return int(gross * 0.4) - 100_000
    elif gross <= 3_600_000:
        return int(gross * 0.3) + 80_000
    elif gross <= 6_600_000:
        return int(gross * 0.2) + 440_000
    elif gross <= 8_500_000:
        return int(gross * 0.1) + 1_100_000
    else:
        return 1_950_000

def _life_ins_deduction(paid):
    """生命保険料控除（新契約・一般）"""
    if paid <= 0:
        return 0
    elif paid <= 20_000:
        return paid
    elif paid <= 40_000:
        return int(paid * 0.5) + 10_000
    elif paid <= 80_000:
        return int(paid * 0.25) + 20_000
    else:
        return 40_000

def _income_tax(taxable):
    """所得税額（超過累進課税）"""
    if taxable <= 0:
        return 0
    brackets = [
        (1_950_000, 0.05, 0),
        (3_300_000, 0.10, 97_500),
        (6_950_000, 0.20, 427_500),
        (9_000_000, 0.23, 636_000),
        (18_000_000, 0.33, 1_536_000),
        (40_000_000, 0.40, 2_796_000),
    ]
    for limit, rate, deduct in brackets:
        if taxable <= limit:
            return int(taxable * rate) - deduct
    return int(taxable * 0.45) - 4_796_000

def _spouse_deduction(spouse_income):
    """配偶者控除額（配偶者の合計所得が0の場合38万円）"""
    if spouse_income <= 480_000:
        return 380_000
    elif spouse_income <= 1_330_000:
        # 配偶者特別控除（簡易版）
        return max(0, 380_000 - max(0, int((spouse_income - 480_000) / 10_000) * 10_000 // 50_000 * 30_000))
    else:
        return 0

def _calc_year_end(name, year, adj):
    """年末調整メイン計算。結果をdictで返す。"""
    emp = query("SELECT * FROM employees WHERE name=?", (name,))
    if not emp:
        return None
    emp = dict(emp[0])

    rows = query(
        "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? ORDER BY clock_in",
        (name, f"{year}-01-01", f"{year}-12-31")
    )

    gross = 0
    total_transport = 0
    work_days = 0
    for r in rows:
        r = dict(r)
        if not r["clock_out"]:
            continue
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S")
        bm = r.get("break_min") or 0
        ci_m = ceil15(ci.hour * 60 + ci.minute)
        co_m = floor15(co.hour * 60 + co.minute)
        total = max(0, co_m - ci_m - bm)
        std  = min(total, STANDARD_HOURS * 60)
        over = floor15(max(0, total - STANDARD_HOURS * 60))

        emp_at = get_emp_at(name, ci.strftime("%Y-%m-%d"))
        hourly_at    = int(emp_at["hourly_rate"] or 0) if emp_at else 0
        transport_at = int(emp_at["transport_allowance"] or 0) if emp_at else 0
        other_at     = int(emp_at["other_allowance"] or 0) if emp_at else 0

        gross += int(hourly_at * std / 60) + int(hourly_at * over / 60) + other_at
        total_transport += transport_at
        work_days += 1

    # 交通費は月10万円まで非課税（全額非課税扱い）
    taxable_gross = gross

    # 給与所得控除
    emp_ded   = _emp_deduction(taxable_gross)
    emp_inc   = max(0, taxable_gross - emp_ded)

    # 各種所得控除
    basic          = 480_000
    social         = int(adj.get("social_insurance") or 0)
    life_paid      = int(adj.get("life_insurance") or 0)
    life_ded       = _life_ins_deduction(life_paid)
    eq_paid        = int(adj.get("earthquake_insurance") or 0)
    eq_ded         = min(eq_paid, 50_000)
    disability_ded = 270_000 if int(adj.get("disability") or 0) else 0
    dep_ded        = (
        int(adj.get("dependent_count") or 0) * 380_000 +
        int(adj.get("specific_dependent") or 0) * 630_000 +
        int(adj.get("old_dependent") or 0) * 480_000
    )
    spouse_income  = int(adj.get("spouse_income") or 0)
    spouse_ded     = _spouse_deduction(spouse_income) if spouse_income >= 0 else 0

    total_ded = basic + social + life_ded + eq_ded + disability_ded + dep_ded + spouse_ded

    # 課税所得（1000円未満切り捨て）
    taxable = max(0, emp_inc - total_ded)
    taxable = (taxable // 1000) * 1000

    # 所得税 + 復興特別所得税（100円未満切り捨て）
    inc_tax  = _income_tax(taxable)
    fukkou   = int(inc_tax * 0.021)
    year_tax = ((inc_tax + fukkou) // 100) * 100

    withheld = int(adj.get("withheld_tax") or 0)
    diff     = withheld - year_tax  # +還付 / -追徴

    return dict(
        name=name, year=year,
        gross=taxable_gross, transport=total_transport, work_days=work_days,
        emp_ded=emp_ded, emp_inc=emp_inc,
        basic=basic, social=social,
        life_paid=life_paid, life_ded=life_ded,
        eq_paid=eq_paid, eq_ded=eq_ded,
        disability_ded=disability_ded,
        dep_ded=dep_ded, spouse_ded=spouse_ded,
        total_ded=total_ded,
        taxable=taxable,
        inc_tax=inc_tax, fukkou=fukkou, year_tax=year_tax,
        withheld=withheld, diff=diff,
    )


def _make_gensen_excel(result, adj):
    """源泉徴収票 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "源泉徴収票"

    thin = Side(style="thin")
    med  = Side(style="medium")
    def border(t=None,b=None,l=None,r=None):
        return Border(top=t,bottom=b,left=l,right=r)

    ws.page_setup.paperSize  = 9   # A4
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left  = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top   = 0.6
    ws.page_margins.bottom = 0.6

    # 列幅
    for col, w in enumerate([3,12,12,12,12,14,12,12,12,12], 1):
        ws.column_dimensions[chr(64+col)].width = w
    for r in range(1, 30):
        ws.row_dimensions[r].height = 18

    title_font = Font(name="MS明朝", size=14, bold=True)
    hdr_font   = Font(name="MS明朝", size=8)
    val_font   = Font(name="MS明朝", size=10)
    sub_font   = Font(name="MS明朝", size=7)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    def c(row, col, value="", font=None, align=None, fill=None, num_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:      cell.font       = font
        if align:     cell.alignment  = align
        if fill:      cell.fill       = fill
        if num_format:cell.number_format = num_format
        return cell

    header_fill = PatternFill("solid", fgColor="C8A878")
    light_fill  = PatternFill("solid", fgColor="FAF5EE")

    # タイトル
    ws.merge_cells("A1:J1")
    c(1,1, f"給与所得の源泉徴収票　{result['year']}年分（令和{result['year']-2018}年分）",
      font=title_font, align=center)

    # 支払を受ける者
    ws.merge_cells("A3:B3")
    c(3,1, "支払を受ける者", font=hdr_font, align=center)
    ws.merge_cells("C3:J3")
    c(3,3, result["name"], font=val_font, align=left)

    # ヘッダー行
    headers = [
        ("支払金額","gross"),
        ("給与所得控除後\nの金額","emp_inc"),
        ("所得控除の額\nの合計額","total_ded"),
        ("源泉徴収税額","withheld"),
        ("確定年税額","year_tax"),
        ("差引過不足\n（還付＋/追徴－）","diff"),
    ]
    col_start = 2
    for i,(hdr,_) in enumerate(headers):
        col = col_start + i
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col)
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.font = Font(name="MS明朝", size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        ws.row_dimensions[5].height = 28

    # 値行
    for i,(_,key) in enumerate(headers):
        col = col_start + i
        val = result[key]
        cell = ws.cell(row=6, column=col, value=val)
        cell.font = val_font
        cell.alignment = right
        cell.number_format = "#,##0"
        cell.fill = light_fill
        if key == "diff":
            cell.font = Font(name="MS明朝", size=10,
                             color="C03030" if val < 0 else "3A7A3A",
                             bold=True)

    ws.row_dimensions[6].height = 22

    # 控除明細
    detail = [
        ("基礎控除",          result["basic"]),
        ("社会保険料控除",    result["social"]),
        ("生命保険料控除",    result["life_ded"]),
        ("地震保険料控除",    result["eq_ded"]),
        ("扶養控除",          result["dep_ded"]),
        ("配偶者控除",        result["spouse_ded"]),
        ("障害者控除",        result["disability_ded"]),
    ]
    c(8, 1, "控除明細", font=Font(name="MS明朝",size=9,bold=True), align=center)
    for i,(lbl,val) in enumerate(detail):
        row = 9 + i
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c(row,1, lbl, font=hdr_font, align=left)
        c(row,3, val, font=val_font, align=right, num_format="#,##0")

    # 課税所得
    c(17,1, "課税所得", font=Font(name="MS明朝",size=9,bold=True), align=left)
    c(17,3, result["taxable"], font=val_font, align=right, num_format="#,##0")
    c(18,1, "所得税", font=hdr_font, align=left)
    c(18,3, result["inc_tax"], font=val_font, align=right, num_format="#,##0")
    c(19,1, "復興特別所得税(2.1%)", font=hdr_font, align=left)
    c(19,3, result["fukkou"], font=val_font, align=right, num_format="#,##0")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name_safe = result["name"].replace(" ","_")
    return send_file(buf, as_attachment=True,
                     download_name=f"源泉徴収票_{name_safe}_{result['year']}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/yearend", methods=["GET"])
@admin_required
def admin_yearend():
    year = int(request.args.get("year", datetime.now(JST).year))
    staff_rows = query("SELECT name FROM staff ORDER BY name")
    items = []
    for s in staff_rows:
        name = s["name"]
        rows = query("SELECT * FROM year_end_adj WHERE name=? AND year=?", (name, year))
        adj  = dict(rows[0]) if rows else {}
        calc = _calc_year_end(name, year, adj) if rows else None
        items.append(dict(name=name, done=bool(rows), calc=calc))
    return render_template("yearend.html", year=year, items=items,
                           years=list(range(datetime.now(JST).year, 2022, -1)))


@app.route("/admin/yearend/<int:year>/<name>", methods=["GET", "POST"])
@admin_required
def admin_yearend_detail(year, name):
    if request.method == "POST":
        fields = {
            "social_insurance":     int(request.form.get("social_insurance") or 0),
            "life_insurance":       int(request.form.get("life_insurance") or 0),
            "earthquake_insurance": int(request.form.get("earthquake_insurance") or 0),
            "disability":           1 if request.form.get("disability") else 0,
            "dependent_count":      int(request.form.get("dependent_count") or 0),
            "specific_dependent":   int(request.form.get("specific_dependent") or 0),
            "old_dependent":        int(request.form.get("old_dependent") or 0),
            "spouse_income":        int(request.form.get("spouse_income") or -1),
            "withheld_tax":         int(request.form.get("withheld_tax") or 0),
        }
        fields["has_spouse"] = 1 if fields["spouse_income"] >= 0 else 0

        existing = query("SELECT id FROM year_end_adj WHERE name=? AND year=?", (name, year))
        p = ph()
        if existing:
            execute("""UPDATE year_end_adj SET
                social_insurance=?, life_insurance=?, earthquake_insurance=?,
                disability=?, dependent_count=?, specific_dependent=?, old_dependent=?,
                spouse_income=?, withheld_tax=?
                WHERE name=? AND year=?""".replace("?", p),
                (fields["social_insurance"], fields["life_insurance"], fields["earthquake_insurance"],
                 fields["disability"], fields["dependent_count"], fields["specific_dependent"],
                 fields["old_dependent"], fields["spouse_income"], fields["withheld_tax"],
                 name, year))
        else:
            execute("""INSERT INTO year_end_adj
                (name, year, social_insurance, life_insurance, earthquake_insurance,
                 disability, dependent_count, specific_dependent, old_dependent,
                 spouse_income, withheld_tax)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""".replace("?", p),
                (name, year, fields["social_insurance"], fields["life_insurance"],
                 fields["earthquake_insurance"], fields["disability"],
                 fields["dependent_count"], fields["specific_dependent"],
                 fields["old_dependent"], fields["spouse_income"], fields["withheld_tax"]))

        if request.form.get("dl_excel"):
            adj = fields
            result = _calc_year_end(name, year, adj)
            if result:
                return _make_gensen_excel(result, adj)

        return redirect(url_for("admin_yearend_detail", year=year, name=name))

    rows = query("SELECT * FROM year_end_adj WHERE name=? AND year=?", (name, year))
    adj  = dict(rows[0]) if rows else {}
    result = _calc_year_end(name, year, adj) if rows else None
    return render_template("yearend_detail.html", year=year, name=name, adj=adj, result=result)


# ── 給与明細 Excel 出力 ───────────────────────────────────────────

def _save_yearend_adj(name, year, form):
    """フォームから年末調整データを保存し adjdict を返す"""
    fields = {
        "social_insurance":     int(form.get("social_insurance") or 0),
        "life_insurance":       int(form.get("life_insurance") or 0),
        "earthquake_insurance": int(form.get("earthquake_insurance") or 0),
        "disability":           1 if form.get("disability") else 0,
        "dependent_count":      int(form.get("dependent_count") or 0),
        "specific_dependent":   int(form.get("specific_dependent") or 0),
        "old_dependent":        int(form.get("old_dependent") or 0),
        "spouse_income":        int(form.get("spouse_income")) if form.get("spouse_income") != "" else -1,
        "withheld_tax":         int(form.get("withheld_tax") or 0),
    }
    existing = query("SELECT id FROM year_end_adj WHERE name=? AND year=?", (name, year))
    p = ph()
    if existing:
        execute("""UPDATE year_end_adj SET
            social_insurance=?, life_insurance=?, earthquake_insurance=?,
            disability=?, dependent_count=?, specific_dependent=?, old_dependent=?,
            spouse_income=?, withheld_tax=?
            WHERE name=? AND year=?""".replace("?", p),
            (fields["social_insurance"], fields["life_insurance"], fields["earthquake_insurance"],
             fields["disability"], fields["dependent_count"], fields["specific_dependent"],
             fields["old_dependent"], fields["spouse_income"], fields["withheld_tax"],
             name, year))
    else:
        execute("""INSERT INTO year_end_adj
            (name, year, social_insurance, life_insurance, earthquake_insurance,
             disability, dependent_count, specific_dependent, old_dependent,
             spouse_income, withheld_tax)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""".replace("?", p),
            (name, year, fields["social_insurance"], fields["life_insurance"],
             fields["earthquake_insurance"], fields["disability"],
             fields["dependent_count"], fields["specific_dependent"],
             fields["old_dependent"], fields["spouse_income"], fields["withheld_tax"]))
    return fields


@app.route("/admin/payslip", methods=["GET", "POST"])
@admin_required
def admin_payslip():
    staff_rows = query("SELECT name FROM staff ORDER BY name")
    default_start, default_end = current_pay_period()
    date_from = request.args.get("from", default_start.isoformat())
    date_to   = request.args.get("to",   default_end.isoformat())
    target    = request.args.get("name", "")
    download  = request.args.get("dl", "")

    # 年末調整 POST（控除保存 or Excel出力）
    if request.method == "POST" and target:
        calc_year = int(date_from[:4])
        adj = _save_yearend_adj(target, calc_year, request.form)
        if request.form.get("dl_gensen"):
            yr = _calc_year_end(target, calc_year, adj)
            if yr:
                return _make_gensen_excel(yr, adj)
        return redirect(url_for("admin_payslip", name=target,
                                **{"from": date_from, "to": date_to}))

    results = []
    if target:
        rows = query(
            "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? ORDER BY clock_in",
            (target, date_from, date_to)
        )
        for r in rows:
            ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
            co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S") if r["clock_out"] else None
            break_min = r["break_min"] or 0
            if co:
                ci_min = ceil15(ci.hour * 60 + ci.minute)
                co_min = floor15(co.hour * 60 + co.minute)
                total  = max(0, co_min - ci_min - break_min)
                std    = min(total, STANDARD_HOURS * 60)
                over   = floor15(max(0, total - STANDARD_HOURS * 60))
            else:
                std = over = None
            results.append({
                "id": r["id"],
                "date": ci.strftime("%Y/%m/%d"),
                "clock_in": ci.strftime("%H:%M"),
                "clock_out": co.strftime("%H:%M") if co else "—",
                "break": fmt_time(break_min) if break_min else "—",
                "work": fmt_time(std) if std is not None else "出勤中",
                "overtime": fmt_time(over) if over else "—",
                "work_min": std or 0,
                "over_min": over or 0,
            })

        if download == "1" and results:
            emp = query("SELECT * FROM employees WHERE name=?", (target,))
            emp = emp[0] if emp else {"hourly_rate": 0, "transport_allowance": 0, "other_allowance": 0}
            return _make_payslip_excel(target, date_from, date_to, results, emp)

    # 年末調整データ
    calc_year = int(date_from[:4])
    ye_rows = query("SELECT * FROM year_end_adj WHERE name=? AND year=?", (target, calc_year)) if target else []
    ye_adj  = dict(ye_rows[0]) if ye_rows else {}
    ye_result = _calc_year_end(target, calc_year, ye_adj) if (target and ye_rows) else None

    return render_template("payslip.html",
        staff_rows=staff_rows, target=target,
        date_from=date_from, date_to=date_to,
        results=results,
        period_label=pay_period_label(date.fromisoformat(date_from), date.fromisoformat(date_to)),
        total_work=fmt_time(sum(r["work_min"] for r in results)),
        total_over=fmt_time(sum(r["over_min"] for r in results)),
        days=len(results),
        calc_year=calc_year,
        ye_adj=ye_adj,
        ye_result=ye_result)

@app.route("/admin/payslip/pdf")
@admin_required
def admin_payslip_pdf():
    target    = request.args.get("name", "")
    default_start, default_end = current_pay_period()
    date_from = request.args.get("from", default_start.isoformat())
    date_to   = request.args.get("to",   default_end.isoformat())
    if not target:
        return redirect(url_for("admin_payslip"))

    rows = query(
        "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? AND clock_out IS NOT NULL ORDER BY clock_in",
        (target, date_from, date_to)
    )
    total_min = over_min = 0
    for r in rows:
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S")
        break_min = r["break_min"] or 0
        ci_min = ceil15(ci.hour * 60 + ci.minute)
        co_min = floor15(co.hour * 60 + co.minute)
        total  = max(0, co_min - ci_min - break_min)
        std    = min(total, STANDARD_HOURS * 60)
        over   = floor15(max(0, total - STANDARD_HOURS * 60))
        total_min += std + over
        over_min  += over

    emp = query("SELECT * FROM employees WHERE name=?", (target,))
    emp = emp[0] if emp else {"hourly_rate": 0, "transport_allowance": 0, "other_allowance": 0}
    hourly    = int(emp["hourly_rate"] or 0)
    transport = int(emp["transport_allowance"] or 0)
    other_all = int(emp["other_allowance"] or 0)

    reg_min      = total_min - over_min
    base_pay     = int(hourly * reg_min / 60)
    overtime_pay = int(hourly * over_min / 60)
    gross_pay    = base_pay + overtime_pay + transport + other_all

    df_start = date.fromisoformat(date_from)
    df_end   = date.fromisoformat(date_to)
    slip = {
        "name": target,
        "reiwa_year": df_start.year - 2018,
        "month": df_start.month,
        "period_label": pay_period_label(df_start, df_end),
        "pay_date": f"{df_end.year}年{df_end.month}月25日",
        "days": len(rows),
        "total_work": fmt_time(total_min),
        "total_over": fmt_time(over_min) if over_min else "0:00",
        "hourly": hourly, "base_pay": base_pay, "overtime_pay": overtime_pay,
        "transport": transport, "other_all": other_all, "gross_pay": gross_pay,
    }
    return _make_payslip_pdf(slip, date_from, date_to)


def _make_payslip_pdf(slip, date_from, date_to):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.lib.units import mm
    except ImportError:
        return "reportlab が未インストールです。", 500

    # 日本語フォント（埋め込みTTF）を登録。失敗時はCIDフォントへフォールバック
    FONT = "JPFont"
    font_path = os.path.join(os.path.dirname(__file__),
                             "static", "fonts", "SawarabiGothic-Regular.ttf")
    try:
        if FONT not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(FONT, font_path))
    except Exception:
        FONT = "HeiseiKakuGo-W5"
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(FONT))
        except Exception:
            FONT = "Helvetica"

    buf = io.BytesIO()
    W, H = A4
    c = canvas.Canvas(buf, pagesize=A4)
    yen = lambda n: "¥{:,}".format(n)

    left, right = 25*mm, W - 25*mm
    y = H - 25*mm

    # タイトル
    c.setFont(FONT, 9)
    c.drawCentredString(W/2, y, "パン屋ニチゲツ")
    y -= 10*mm
    c.setFont(FONT, 18)
    c.drawCentredString(W/2, y, "給 与 明 細 書")
    y -= 8*mm
    c.setFont(FONT, 10)
    c.drawCentredString(W/2, y, f"令和{slip['reiwa_year']}年{slip['month']}月分　（{slip['period_label']}）")
    y -= 4*mm
    c.setLineWidth(1.2)
    c.line(left, y, right, y)
    y -= 10*mm

    # 氏名・支給日
    c.setFont(FONT, 12)
    c.drawString(left, y, f"{slip['name']} 様")
    c.setFont(FONT, 10)
    c.drawRightString(right, y, f"支給日：{slip['pay_date']}")
    y -= 10*mm

    # 勤怠情報
    c.setFont(FONT, 10)
    c.drawString(left, y,
        f"出勤日数：{slip['days']}日　　総労働時間：{slip['total_work']}　　残業時間：{slip['total_over']}")
    y -= 12*mm

    # 支給明細
    def row(label, amount, bold=False, line_above=False):
        nonlocal y
        if line_above:
            c.setLineWidth(0.6)
            c.line(left, y + 4*mm, right, y + 4*mm)
        c.setFont(FONT, 12 if bold else 11)
        c.drawString(left + 4*mm, y, label)
        c.drawRightString(right - 4*mm, y, amount)
        y -= 9*mm

    c.setFont(FONT, 10)
    c.drawString(left, y, "◆ 支給")
    y -= 8*mm
    row("基本給", yen(slip["base_pay"]))
    if slip["overtime_pay"] > 0:
        row("残業手当", yen(slip["overtime_pay"]))
    if slip["transport"] > 0:
        row("通勤手当", yen(slip["transport"]))
    if slip["other_all"] > 0:
        row("その他手当", yen(slip["other_all"]))
    row("総支給額", yen(slip["gross_pay"]), bold=True, line_above=True)
    y -= 6*mm

    # 差引支給額ボックス
    box_h = 16*mm
    c.setFillColorRGB(0.78, 0.47, 0.25)
    c.rect(left, y - box_h + 6*mm, right - left, box_h, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont(FONT, 12)
    c.drawString(left + 6*mm, y - box_h + 12*mm, "差引支給額")
    c.setFont(FONT, 16)
    c.drawRightString(right - 6*mm, y - box_h + 11*mm, yen(slip["gross_pay"]))
    c.setFillColorRGB(0, 0, 0)

    # フッター
    c.setFont(FONT, 8)
    c.setFillColorRGB(0.6, 0.6, 0.6)
    c.drawRightString(right, 20*mm, "発行: パン屋ニチゲツ")

    c.showPage()
    c.save()
    buf.seek(0)
    filename = f"payslip_{slip['name']}_{date_from}_{date_to}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/pdf")

@app.route("/admin/payslip/all_excel")
@admin_required
def admin_payslip_all_excel():
    default_start, default_end = current_pay_period()
    date_from = request.args.get("from", default_start.isoformat())
    date_to   = request.args.get("to",   default_end.isoformat())

    staff_rows = query("SELECT name FROM staff ORDER BY name")
    payslip_data = []

    for s in staff_rows:
        name = s["name"]
        rows = query(
            "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? AND clock_out IS NOT NULL ORDER BY clock_in",
            (name, date_from, date_to)
        )
        results = []
        for r in rows:
            ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
            co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S")
            break_min = r["break_min"] or 0
            ci_min = ceil15(ci.hour * 60 + ci.minute)
            co_min = floor15(co.hour * 60 + co.minute)
            total  = max(0, co_min - ci_min - break_min)
            std    = min(total, STANDARD_HOURS * 60)
            over   = floor15(max(0, total - STANDARD_HOURS * 60))
            results.append({"work_min": std, "over_min": over})

        emp = query("SELECT * FROM employees WHERE name=?", (name,))
        emp = emp[0] if emp else {"hourly_rate": 0, "transport_allowance": 0, "other_allowance": 0}
        payslip_data.append({"name": name, "results": results, "emp": emp})

    return _make_all_payslip_excel(payslip_data, date_from, date_to)

def _write_one_slip_row(ws, row_offset, name, days,
                        total_h, total_m, over_h, over_m, hourly,
                        base_pay, overtime_pay, transport,
                        gross_pay,
                        reiwa_year, month, df_start, df_end,
                        Font, Alignment, Border, Side, PatternFill):
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left",   vertical="center")
    ra = Alignment(horizontal="right",  vertical="center")
    gray = PatternFill("solid", fgColor="DDDDDD")
    blue = PatternFill("solid", fgColor="B8CCE4")

    NCOLS = 10  # 総列数

    def cell(ro, co, val="", bold=False, align=ca, border=False, fill=None, size=11):
        c = ws.cell(row=row_offset + ro, column=co, value=val)
        c.font = Font(name="MS明朝", bold=bold, size=size)
        c.alignment = align
        if border:
            c.border = thin
        if fill:
            c.fill = fill

    def mmerge(ro, c1, c2):
        ws.merge_cells(start_row=row_offset+ro, start_column=c1,
                       end_row=row_offset+ro, end_column=c2)

    # ── タイトル行 (row 0) ──
    mmerge(0, 1, 2);  cell(0, 1, "パート給与明細書", bold=True, align=ca, size=12)
    mmerge(0, 3, 4);  cell(0, 3, f"令和{reiwa_year}年{month}月分", align=ca)
    mmerge(0, 5, 7);  cell(0, 5, f"氏名：{name}", align=la)
    mmerge(0, 8, 9);  cell(0, 8,
        f"対象期間：{df_start.month}/{df_start.day}〜{df_end.month}/{df_end.day}",
        align=ca)
    mmerge(0, 10, NCOLS); cell(0, 10,
        f"支給日：{df_end.year}年{df_end.month}月25日",
        align=ra)

    # ── ヘッダー行 (row 1) ──
    headers = ["出勤日数", "労働時間", "残業時間", "時給",
               "基本給", "通勤手当", "総支給額",
               "所得税", "控除合計", "差引支給額"]
    for i, h in enumerate(headers):
        col = i + 1
        fill = blue if i == len(headers) - 1 else gray
        cell(1, col, h, align=ca, border=True, fill=fill)

    # ── データ行 (row 2) ──
    values = [
        (f"{days}日",                   ca),
        (f"{total_h}:{total_m:02d}",    ca),
        (f"{over_h}:{over_m:02d}",      ca),
        (f"{hourly:,}",                 ra),
        (f"{base_pay + overtime_pay:,}", ra),
        (f"{transport:,}",              ra),
        (f"{gross_pay:,}",              ra),
        ("",                            ra),  # 所得税は手入力
        ("",                            ra),  # 控除合計は手入力
        (f"{gross_pay:,}",              ra),
    ]
    for i, (v, a) in enumerate(values):
        col = i + 1
        is_net = (i == len(values) - 1)
        cell(2, col, v, bold=is_net, align=a, border=True)


def _make_all_payslip_excel(payslip_data, date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl が未インストールです。", 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "給与明細"

    # A4横（landscape）設定
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True

    # 列幅（10列）
    col_widths = [9, 10, 10, 9, 13, 11, 13, 10, 11, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    df_start   = date.fromisoformat(date_from)
    df_end     = date.fromisoformat(date_to)
    reiwa_year = df_start.year - 2018
    month      = df_start.month

    SLIP_ROWS = 3  # タイトル行 + ヘッダー行 + データ行
    GAP_ROW   = 1

    for idx, pdata in enumerate(payslip_data):
        row_off = 1 + idx * (SLIP_ROWS + GAP_ROW)
        # 行高
        for r in range(row_off, row_off + SLIP_ROWS):
            ws.row_dimensions[r].height = 22

        name      = pdata["name"]
        results   = pdata["results"]
        emp       = pdata["emp"]
        hourly    = int(emp["hourly_rate"] or 0)
        transport = int(emp["transport_allowance"] or 0)
        other_all = int(emp["other_allowance"] or 0)
        days      = len(results)

        total_min    = sum(r["work_min"] for r in results)
        over_min     = sum(r["over_min"] for r in results)
        reg_min      = total_min - over_min
        total_h, total_m = divmod(total_min, 60)
        over_h,  over_m  = divmod(over_min,  60)
        base_pay     = int(hourly * reg_min / 60)
        overtime_pay = int(hourly * over_min / 60)
        gross_pay    = base_pay + overtime_pay + transport + other_all

        _write_one_slip_row(ws, row_off, name, days,
                            total_h, total_m, over_h, over_m, hourly,
                            base_pay, overtime_pay, transport,
                            gross_pay,
                            reiwa_year, month, df_start, df_end,
                            Font, Alignment, Border, Side, PatternFill)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"payslips_all_{date_from}_{date_to}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _make_payslip_excel(name, date_from, date_to, results, emp=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return "openpyxl が未インストールです。", 500

    emp = emp or {"hourly_rate": 0, "transport_allowance": 0, "other_allowance": 0}
    hourly    = emp["hourly_rate"]
    transport = emp["transport_allowance"]
    other_all = emp["other_allowance"]

    total_min  = sum(r["work_min"] for r in results)
    over_min   = sum(r["over_min"] for r in results)
    reg_min    = total_min - over_min
    total_h    = total_min // 60
    total_m    = total_min % 60
    days       = len(results)

    base_pay     = int(hourly * reg_min / 60)
    overtime_pay = int(hourly * over_min / 60)
    gross_pay    = base_pay + overtime_pay + transport + other_all

    # 令和年計算
    reiwa_year = date.fromisoformat(date_from).year - 2018
    month      = date.fromisoformat(date_from).month
    df_start   = date.fromisoformat(date_from)
    df_end     = date.fromisoformat(date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "給料支払明細書"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    def cell(r, c, val="", bold=False, size=11, merge=None, align=center, border=None):
        ws.cell(row=r, column=c, value=val)
        ws.cell(row=r, column=c).font = Font(name="MS明朝", bold=bold, size=size)
        ws.cell(row=r, column=c).alignment = align
        if merge:
            ws.merge_cells(merge)
        if border:
            ws.cell(row=r, column=c).border = thin

    # 列幅
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 2

    # 行高
    for i in range(1, 20):
        ws.row_dimensions[i].height = 20

    # タイトル
    cell(1, 2, "給料支払明細書", bold=True, size=14, merge="B1:D1")
    cell(2, 2, f"（令和{reiwa_year}年{month}月分）", size=11, merge="B2:D2")
    cell(3, 2, f"{name}様", size=11, merge="B3:D3")
    cell(4, 2, "")  # 空行
    cell(5, 2, f"労働日数　自{df_start.month}月{df_start.day}日　至{df_end.month}月{df_end.day}日",
         size=10, merge="B5:D5", border=True, align=left)

    # 支給額ヘッダー
    ws.merge_cells("B7:D7")
    ws["B7"] = "支給額"
    ws["B7"].font = Font(name="MS明朝", size=10)
    ws["B7"].alignment = center
    ws["B7"].border = thin
    ws["C7"].border = thin
    ws["D7"].border = thin

    # 労働時間・金額行
    ws["B8"] = f"{total_h}時間　{total_m:02d}分"
    ws["B8"].font = Font(name="MS明朝", size=10)
    ws["B8"].alignment = center
    ws["B8"].border = thin
    ws["C8"].border = thin
    ws["D8"] = f"{base_pay + overtime_pay:,}円"
    ws["D8"].font = Font(name="MS明朝", size=10)
    ws["D8"].alignment = right
    ws["D8"].border = thin

    # 交通費
    ws["B9"] = "交通費"
    ws["B9"].font = Font(name="MS明朝", size=10)
    ws["B9"].alignment = left
    ws["B9"].border = thin
    ws["C9"].border = thin
    ws["D9"] = f"{transport:,}円" if transport else "円"
    ws["D9"].font = Font(name="MS明朝", size=10)
    ws["D9"].alignment = right
    ws["D9"].border = thin

    # 合計
    ws["B10"] = "合計"
    ws["B10"].font = Font(name="MS明朝", size=10, bold=True)
    ws["B10"].alignment = left
    ws["B10"].border = thin
    ws["C10"].border = thin
    ws["D10"] = f"{gross_pay:,}円"
    ws["D10"].font = Font(name="MS明朝", size=10, bold=True)
    ws["D10"].alignment = right
    ws["D10"].border = thin

    # 控除額ヘッダー
    ws.merge_cells("B11:D11")
    ws["B11"] = "控除額"
    ws["B11"].font = Font(name="MS明朝", size=10)
    ws["B11"].alignment = center
    ws["B11"].border = thin
    ws["C11"].border = thin
    ws["D11"].border = thin

    # 所得税
    ws["B12"] = "所得税"
    ws["B12"].font = Font(name="MS明朝", size=10)
    ws["B12"].alignment = left
    ws["B12"].border = thin
    ws["C12"].border = thin
    ws["D12"] = "円"
    ws["D12"].font = Font(name="MS明朝", size=10)
    ws["D12"].alignment = right
    ws["D12"].border = thin

    # 空白控除行
    ws["B13"].border = thin
    ws["C13"].border = thin
    ws["D13"] = "円"
    ws["D13"].font = Font(name="MS明朝", size=10)
    ws["D13"].alignment = right
    ws["D13"].border = thin

    # 空白控除行
    ws["B14"].border = thin
    ws["C14"].border = thin
    ws["D14"] = "円"
    ws["D14"].font = Font(name="MS明朝", size=10)
    ws["D14"].alignment = right
    ws["D14"].border = thin

    # 控除合計
    ws["B15"] = "合計"
    ws["B15"].font = Font(name="MS明朝", size=10)
    ws["B15"].alignment = left
    ws["B15"].border = thin
    ws["C15"].border = thin
    ws["D15"] = "円"
    ws["D15"].font = Font(name="MS明朝", size=10)
    ws["D15"].alignment = right
    ws["D15"].border = thin

    # 差引支給額
    ws["B16"] = "差引支給額"
    ws["B16"].font = Font(name="MS明朝", bold=True, size=10)
    ws["B16"].alignment = left
    ws["B16"].border = thin
    ws["C16"].border = thin
    ws["D16"] = f"{gross_pay:,}円"
    ws["D16"].font = Font(name="MS明朝", bold=True, size=10)
    ws["D16"].alignment = right
    ws["D16"].border = thin

    # 会社名
    ws["D17"] = "ぱんやニチゲツ"
    ws["D17"].font = Font(name="MS明朝", size=9)
    ws["D17"].alignment = right

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"payslip_{name}_{date_from}_{date_to}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── 日次売上 ─────────────────────────────────────────────────────

@app.route("/admin/sales", methods=["GET", "POST"])
@admin_required
def admin_sales():
    error = msg = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            sale_date = request.form.get("date", today_jst())
            amount    = request.form.get("amount", "").strip()
            category  = request.form.get("category", "").strip()
            memo      = request.form.get("memo", "").strip()
            if not amount.isdigit():
                error = "金額は数字で入力してください"
            else:
                execute("INSERT INTO sales (date, amount, category, memo) VALUES (?, ?, ?, ?)",
                        (sale_date, int(amount), category, memo))
                msg = "売上を記録しました ✓"
        elif action == "delete":
            sale_id = request.form.get("sale_id")
            if sale_id:
                execute("DELETE FROM sales WHERE id=?", (sale_id,))

    month = request.args.get("month", today_jst()[:7])
    rows  = query("SELECT * FROM sales WHERE date LIKE ? ORDER BY date DESC", (f"{month}%",))
    total = sum(r["amount"] for r in rows)

    monthly = []
    seen = set()
    all_rows = query("SELECT DISTINCT substr(date,1,7) as m FROM sales ORDER BY m DESC")
    for r in all_rows:
        monthly.append(r["m"])

    return render_template("sales.html",
        rows=rows, total=total, month=month,
        monthly=monthly, today=today_jst(),
        msg=msg, error=error)

# ── シフト管理 ────────────────────────────────────────────────────

@app.route("/admin/shifts", methods=["GET", "POST"])
@admin_required
def admin_shifts():
    msg = error = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            name       = request.form.get("name", "").strip()
            shift_date = request.form.get("date", "").strip()
            start_time = request.form.get("start_time", "").strip()
            end_time   = request.form.get("end_time", "").strip()
            if not all([name, shift_date, start_time, end_time]):
                error = "全項目を入力してください"
            else:
                execute("INSERT INTO shifts (name, date, start_time, end_time) VALUES (?, ?, ?, ?)",
                        (name, shift_date, start_time, end_time))
                msg = "シフトを登録しました ✓"
        elif action == "delete":
            shift_id = request.form.get("shift_id")
            if shift_id:
                execute("DELETE FROM shifts WHERE id=?", (shift_id,))
                msg = "削除しました"

    # 表示週（月曜始まり）
    week_str = request.args.get("week", "")
    if week_str:
        try:
            week_start = date.fromisoformat(week_str)
        except ValueError:
            week_start = date.today()
    else:
        week_start = date.today()
    # 月曜に揃える
    week_start = week_start - timedelta(days=week_start.weekday())
    week_end   = week_start + timedelta(days=6)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]

    rows = query(
        "SELECT * FROM shifts WHERE date BETWEEN ? AND ? ORDER BY date, start_time",
        (week_start.isoformat(), week_end.isoformat())
    )
    staff_rows = query("SELECT name FROM staff ORDER BY name")

    # 週別グリッド: {name: {date_str: [shifts]}}
    grid = defaultdict(lambda: defaultdict(list))
    for r in rows:
        grid[r["name"]][r["date"]].append(r)

    prev_week = (week_start - timedelta(days=7)).isoformat()
    next_week = (week_start + timedelta(days=7)).isoformat()

    return render_template("shifts.html",
        rows=rows, staff_rows=staff_rows,
        week_dates=week_dates, grid=grid,
        prev_week=prev_week, next_week=next_week,
        week_start=week_start,
        today_str=today_jst(),
        msg=msg, error=error)

# ── 商品・レシピ管理 ──────────────────────────────────────────────

@app.route("/admin/products", methods=["GET", "POST"])
@admin_required
def admin_products():
    msg = error = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            name     = request.form.get("name", "").strip()
            category = request.form.get("category", "").strip()
            price    = request.form.get("price", "0").strip()
            cost     = request.form.get("cost", "0").strip()
            recipe   = request.form.get("recipe", "").strip()
            if not name:
                error = "商品名を入力してください"
            else:
                try:
                    execute("INSERT INTO products (name, category, price, cost, recipe) VALUES (?, ?, ?, ?, ?)",
                            (name, category, int(price or 0), int(cost or 0), recipe))
                    msg = f"「{name}」を登録しました ✓"
                except Exception:
                    error = f"「{name}」はすでに登録されています"
        elif action == "delete":
            prod_id = request.form.get("prod_id")
            if prod_id:
                execute("DELETE FROM products WHERE id=?", (prod_id,))
                msg = "削除しました"
        elif action == "edit":
            prod_id  = request.form.get("prod_id")
            name     = request.form.get("name", "").strip()
            category = request.form.get("category", "").strip()
            price    = request.form.get("price", "0").strip()
            cost     = request.form.get("cost", "0").strip()
            recipe   = request.form.get("recipe", "").strip()
            if prod_id and name:
                execute("UPDATE products SET name=?, category=?, price=?, cost=?, recipe=? WHERE id=?",
                        (name, category, int(price or 0), int(cost or 0), recipe, prod_id))
                msg = "更新しました ✓"

    category_filter = request.args.get("cat", "")
    if category_filter:
        products = query("SELECT * FROM products WHERE category=? ORDER BY category, name", (category_filter,))
    else:
        products = query("SELECT * FROM products ORDER BY category, name")

    categories = query("SELECT DISTINCT category FROM products WHERE category != '' ORDER BY category")

    return render_template("products.html",
        products=products, categories=categories,
        category_filter=category_filter,
        msg=msg, error=error)

# ── ログイン ───────────────────────────────────────────────────────

@app.route("/admin/instagram", methods=["GET", "POST"])
@admin_required
def admin_instagram():
    results = []
    error = None
    if request.method == "POST":
        files = request.files.getlist("media")
        files = [f for f in files if f and f.filename]
        if not files:
            error = "ファイルを選択してください"
        else:
            try:
                import google.generativeai as genai
                import base64, mimetypes, json, re
                genai.configure(api_key=os.environ.get("GOOGLE_API_KEY", ""))
                model = genai.GenerativeModel("gemini-1.5-flash")

                RULES = (
                    "あなたはパン屋「ニチゲツ」のInstagram運用担当です。"
                    "投稿ルール: 製造風景40%・商品紹介30%・裏側20%・お知らせ10%。"
                    "キャプションは1〜3行、感情を書く、お客様目線。禁止: 長文・自慢・専門用語。"
                    "最重要: 「美味しそう」より「今日行かないと無くなりそう」を感じさせる。"
                    "キャプション末尾には必ず「プレオープン中のため売切れ次第終了となります。」を入れる。"
                )

                image_data_list = []
                for f in files:
                    raw = f.read()
                    mime = f.content_type or mimetypes.guess_type(f.filename)[0] or "image/jpeg"
                    b64 = base64.b64encode(raw).decode()
                    image_data_list.append({
                        "filename": f.filename,
                        "mime": mime,
                        "raw": raw,
                        "b64": b64,
                        "is_video": mime.startswith("video/"),
                    })

                for img in image_data_list:
                    if img["is_video"]:
                        prompt = (
                            f"{RULES}\n\n"
                            f"動画ファイル「{img['filename']}」についてInstagram Reelの編集提案をしてください。\n"
                            "以下のJSON形式のみで回答（```json不要）:\n"
                            '{"type":"製造風景", "caption":"キャプション文(改行は\\n)", "edit":"編集提案(カット順・テキスト・BGM)", "order_reason":"投稿順の理由"}'
                        )
                        resp = model.generate_content(prompt)
                    else:
                        prompt = (
                            f"{RULES}\n\n"
                            "この画像を見てInstagramキャプションと投稿タイプを提案してください。\n"
                            "typeは必ず「製造風景」「商品紹介」「裏側」「お知らせ」のいずれか。\n"
                            "以下のJSON形式のみで回答（```json不要）:\n"
                            '{"type":"商品紹介", "caption":"キャプション文(改行は\\n)", "edit":"投稿のポイント・加工提案", "order_reason":"投稿順の理由"}'
                        )
                        image_part = {"mime_type": img["mime"], "data": img["raw"]}
                        resp = model.generate_content([prompt, image_part])

                    raw_text = resp.text
                    m = re.search(r'\{.*\}', raw_text, re.DOTALL)
                    if m:
                        parsed = json.loads(m.group())
                    else:
                        parsed = {"type":"不明","caption":raw_text,"edit":"","order_reason":""}

                    results.append({
                        "filename": img["filename"],
                        "is_video": img["is_video"],
                        "b64": img["b64"] if not img["is_video"] else "",
                        "mime": img["mime"],
                        **parsed
                    })

                # 投稿順を再整理（製造→商品→裏側→お知らせ）
                TYPE_ORDER = {"製造風景":0,"商品紹介":1,"裏側":2,"お知らせ":3}
                results.sort(key=lambda r: TYPE_ORDER.get(r.get("type","お知らせ"), 9))

            except Exception as e:
                error = f"分析エラー: {e}"

    return render_template("instagram.html", results=results, error=error)

# ── 給与管理（月次明細） ──────────────────────────────────────────

def calc_payslip_data(emp, records_for_period):
    hourly = emp["hourly_rate"]
    daily_regular_min = 8 * 60
    total_work_min = regular_min = overtime_min = 0

    for r in records_for_period:
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S") if r["clock_out"] else None
        if co is None:
            continue
        work_min = int((co - ci).total_seconds() // 60) - (r["break_min"] or 0)
        if work_min <= 0:
            continue
        total_work_min += work_min
        regular_min  += min(work_min, daily_regular_min)
        overtime_min += max(0, work_min - daily_regular_min)

    base_pay     = int(hourly * regular_min / 60)
    overtime_pay = int(hourly * overtime_min / 60)
    gross_pay    = base_pay + overtime_pay + emp["transport_allowance"] + emp["other_allowance"]

    return {
        "total_work_min": total_work_min,
        "regular_min":    regular_min,
        "overtime_min":   overtime_min,
        "base_pay":       base_pay,
        "overtime_pay":   overtime_pay,
        "transport_allowance": emp["transport_allowance"],
        "other_allowance":     emp["other_allowance"],
        "gross_pay":      gross_pay,
    }

@app.route("/admin/payroll")
@admin_required
def admin_payroll():
    default_start, default_end = current_pay_period()
    date_from = request.args.get("from", default_start.isoformat())
    date_to   = request.args.get("to",   default_end.isoformat())

    employees = [dict(e) for e in query("SELECT * FROM employees ORDER BY name")]
    for e in employees:
        e.setdefault("fuyou_target", 0)
        e.setdefault("other_income", 0)
    saved_slips = query(
        "SELECT * FROM payslips WHERE period_start=? AND period_end=? ORDER BY employee_name",
        (date_from, date_to)
    )
    issued_names = {p["employee_name"] for p in saved_slips}

    records = query(
        "SELECT * FROM records WHERE date(clock_in) BETWEEN ? AND ? AND clock_out IS NOT NULL",
        (date_from, date_to)
    )
    summary = {}
    for r in records:
        n = r["name"]
        ci = datetime.strptime(r["clock_in"], "%Y-%m-%d %H:%M:%S")
        co = datetime.strptime(r["clock_out"], "%Y-%m-%d %H:%M:%S")
        work_min = int((co - ci).total_seconds() // 60) - (r["break_min"] or 0)
        if n not in summary:
            summary[n] = {"days": 0, "total_min": 0}
        summary[n]["days"] += 1
        summary[n]["total_min"] += max(0, work_min)

    return render_template("payroll.html",
        employees=employees,
        payslips=saved_slips,
        issued_names=issued_names,
        summary=summary,
        date_from=date_from,
        date_to=date_to,
        period_label=pay_period_label(date.fromisoformat(date_from), date.fromisoformat(date_to))
    )

@app.route("/admin/payroll/employee/save", methods=["POST"])
@admin_required
def admin_employee_save():
    emp_id    = request.form.get("id", "").strip()
    name      = request.form["name"].strip()
    hourly       = int(request.form.get("hourly_rate", 0) or 0)
    transport    = int(request.form.get("transport_allowance", 0) or 0)
    other        = int(request.form.get("other_allowance", 0) or 0)
    other_income = int(request.form.get("other_income", 0) or 0)
    fuyou_category = request.form.get("fuyou_category", "103")
    if fuyou_category not in ("103", "123", "none"):
        fuyou_category = "103"
    notes        = request.form.get("notes", "").strip()

    if emp_id:
        try:
            execute(
                "UPDATE employees SET name=?, hourly_rate=?, transport_allowance=?, other_allowance=?, other_income=?, fuyou_category=?, notes=? WHERE id=?",
                (name, hourly, transport, other, other_income, fuyou_category, notes, emp_id)
            )
        except Exception:
            execute(
                "UPDATE employees SET name=?, hourly_rate=?, transport_allowance=?, other_allowance=?, notes=? WHERE id=?",
                (name, hourly, transport, other, notes, emp_id)
            )
    else:
        try:
            execute(
                "INSERT INTO employees (name, hourly_rate, transport_allowance, other_allowance, other_income, fuyou_category, notes) VALUES (?,?,?,?,?,?,?)",
                (name, hourly, transport, other, other_income, fuyou_category, notes)
            )
        except Exception:
            execute(
                "INSERT INTO employees (name, hourly_rate, transport_allowance, other_allowance, notes) VALUES (?,?,?,?,?)",
                (name, hourly, transport, other, notes)
            )
    record_employee_history(name, hourly, transport, other, other_income, fuyou_category)
    return redirect(url_for("admin_payroll"))

@app.route("/admin/payroll/employee/<int:emp_id>/fuyou", methods=["POST"])
@admin_required
def admin_employee_fuyou(emp_id):
    category = request.form.get("fuyou_category", "103")
    if category not in ("103", "123", "none"):
        category = "103"
    emp = query("SELECT * FROM employees WHERE id=?", (emp_id,))
    if emp:
        e = dict(emp[0])
        execute("UPDATE employees SET fuyou_category=? WHERE id=?", (category, emp_id))
        record_employee_history(
            e["name"], int(e["hourly_rate"]), int(e["transport_allowance"]),
            int(e["other_allowance"]), int(e["other_income"] or 0), category
        )
    return redirect(url_for("admin_payroll"))

@app.route("/admin/payroll/employee/<int:emp_id>/delete", methods=["POST"])
@admin_required
def admin_employee_delete(emp_id):
    execute("DELETE FROM employees WHERE id=?", (emp_id,))
    return redirect(url_for("admin_payroll"))

@app.route("/admin/payroll/preview", methods=["POST"])
@admin_required
def admin_payslip_preview():
    emp_name  = request.json.get("employee_name")
    date_from = request.json.get("date_from")
    date_to   = request.json.get("date_to")

    emps = query("SELECT * FROM employees WHERE name=?", (emp_name,))
    if not emps:
        return jsonify({"error": "従業員が見つかりません"})
    emp = emps[0]

    records = query(
        "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? AND clock_out IS NOT NULL",
        (emp_name, date_from, date_to)
    )
    return jsonify(calc_payslip_data(emp, records))

@app.route("/admin/payroll/generate", methods=["POST"])
@admin_required
def admin_payroll_generate():
    emp_name  = request.form["employee_name"]
    date_from = request.form["date_from"]
    date_to   = request.form["date_to"]
    note      = request.form.get("note", "").strip()

    emps = query("SELECT * FROM employees WHERE name=?", (emp_name,))
    if not emps:
        return redirect(url_for("admin_payroll"))
    emp = emps[0]

    records = query(
        "SELECT * FROM records WHERE name=? AND date(clock_in) BETWEEN ? AND ? AND clock_out IS NOT NULL",
        (emp_name, date_from, date_to)
    )
    data = calc_payslip_data(emp, records)

    health  = int(request.form.get("health_insurance", 0) or 0)
    pension = int(request.form.get("pension", 0) or 0)
    emp_ins = int(request.form.get("employment_insurance", 0) or 0)
    tax     = int(request.form.get("income_tax", 0) or 0)
    other_d = int(request.form.get("other_deduction", 0) or 0)
    total_d = health + pension + emp_ins + tax + other_d
    net_pay = data["gross_pay"] - total_d

    execute(
        "DELETE FROM payslips WHERE employee_name=? AND period_start=? AND period_end=?",
        (emp_name, date_from, date_to)
    )
    execute("""
        INSERT INTO payslips (
            employee_name, period_start, period_end,
            total_work_min, regular_min, overtime_min,
            base_pay, overtime_pay, transport_allowance, other_allowance, gross_pay,
            health_insurance, pension, employment_insurance, income_tax, other_deduction,
            total_deduction, net_pay, note, created_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        emp_name, date_from, date_to,
        data["total_work_min"], data["regular_min"], data["overtime_min"],
        data["base_pay"], data["overtime_pay"],
        data["transport_allowance"], data["other_allowance"], data["gross_pay"],
        health, pension, emp_ins, tax, other_d,
        total_d, net_pay, note, now_jst()
    ))

    slip = query(
        "SELECT id FROM payslips WHERE employee_name=? AND period_start=? AND period_end=? ORDER BY id DESC LIMIT 1",
        (emp_name, date_from, date_to)
    )
    return redirect(url_for("admin_payslip_print", slip_id=slip[0]["id"]))

@app.route("/admin/payroll/slip/<int:slip_id>")
@admin_required
def admin_payslip_print(slip_id):
    slips = query("SELECT * FROM payslips WHERE id=?", (slip_id,))
    if not slips:
        return redirect(url_for("admin_payroll"))
    return render_template("payslip_print.html", slip=slips[0])

@app.route("/admin/payroll/slip/<int:slip_id>/delete", methods=["POST"])
@admin_required
def admin_payslip_slip_delete(slip_id):
    execute("DELETE FROM payslips WHERE id=?", (slip_id,))
    return redirect(url_for("admin_payroll"))

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = None
    if request.method == "POST":
        if request.form["password"] == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect(url_for("admin"))
        error = "パスワードが違います"
    return render_template("login.html", error=error)

@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("index"))

try:
    init_db()
    print("DB初期化成功")
except Exception as e:
    print(f"DB初期化エラー: {e}")
    import traceback; traceback.print_exc()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
