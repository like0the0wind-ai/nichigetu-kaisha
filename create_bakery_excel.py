from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

wb = Workbook()
wb.remove(wb.active)

# ── 共通スタイル ──────────────────────────────────────────────
FONT_NAME = "Arial"

def hdr_cell(ws, row, col, value, bg="2F5597", fg="FFFFFF", bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def data_cell(ws, row, col, value=None, fmt=None, color="000000", bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, color=color, size=10)
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    c.alignment = Alignment(vertical="center")
    return c

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def section_title(ws, row, col_end, title, col_start=1):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", start_color="D9E1F2")
    ws.row_dimensions[row].height = 22

# ── サンプルデータ ────────────────────────────────────────────
PRODUCTS = [
    ("P001", "食パン（1斤）", "食パン", 350, 120, 15, 8),
    ("P002", "ベーグル（プレーン）", "ベーグル", 200, 70, 20, 10),
    ("P003", "クロワッサン", "クロワッサン", 250, 110, 35, 6),
    ("P004", "カレーパン", "惣菜パン", 220, 85, 25, 8),
    ("P005", "メロンパン", "菓子パン", 200, 65, 20, 10),
    ("P006", "チョコデニッシュ", "菓子パン", 230, 90, 30, 6),
    ("P007", "バゲット", "食パン", 400, 130, 25, 5),
    ("P008", "シナモンロール", "焼き菓子", 280, 100, 30, 8),
]

SALES_DATA = [
    (datetime.date(2024, 6, 1), "土", "8-9",  "食パン（1斤）",     3, 350),
    (datetime.date(2024, 6, 1), "土", "8-9",  "クロワッサン",      5, 250),
    (datetime.date(2024, 6, 1), "土", "9-10", "ベーグル（プレーン）", 4, 200),
    (datetime.date(2024, 6, 1), "土", "9-10", "カレーパン",        6, 220),
    (datetime.date(2024, 6, 1), "土", "10-11","メロンパン",        8, 200),
    (datetime.date(2024, 6, 1), "土", "10-11","食パン（1斤）",     2, 350),
    (datetime.date(2024, 6, 1), "土", "11-12","カレーパン",        3, 220),
    (datetime.date(2024, 6, 1), "土", "11-12","クロワッサン",      2, 250),
    (datetime.date(2024, 6, 2), "日", "8-9",  "食パン（1斤）",     4, 350),
    (datetime.date(2024, 6, 2), "日", "9-10", "ベーグル（プレーン）", 3, 200),
    (datetime.date(2024, 6, 2), "日", "9-10", "メロンパン",        5, 200),
    (datetime.date(2024, 6, 2), "日", "10-11","カレーパン",        4, 220),
    (datetime.date(2024, 6, 3), "月", "8-9",  "食パン（1斤）",     2, 350),
    (datetime.date(2024, 6, 3), "月", "9-10", "クロワッサン",      3, 250),
    (datetime.date(2024, 6, 3), "月", "10-11","カレーパン",        2, 220),
]

WASTE_DATA = [
    (datetime.date(2024, 6, 1), "クロワッサン",      2, 250),
    (datetime.date(2024, 6, 1), "メロンパン",        3, 200),
    (datetime.date(2024, 6, 2), "シナモンロール",    4, 280),
    (datetime.date(2024, 6, 2), "チョコデニッシュ",  2, 230),
    (datetime.date(2024, 6, 3), "クロワッサン",      1, 250),
]

LABOR_DATA = [
    (datetime.date(2024, 6, 1), "田中 花子", "06:00", "14:00", 1050),
    (datetime.date(2024, 6, 1), "鈴木 太郎", "08:00", "16:00", 1000),
    (datetime.date(2024, 6, 2), "田中 花子", "06:00", "14:00", 1050),
    (datetime.date(2024, 6, 3), "鈴木 太郎", "07:00", "15:00", 1000),
]

MFG_DATA = [
    ("食パン（1斤）",     "30分", "10分", "50分", 1),
    ("ベーグル（プレーン）","30分", "20分", "25分", 1),
    ("クロワッサン",      "60分", "40分", "20分", 2),
    ("カレーパン",        "30分", "30分", "20分", 1),
    ("メロンパン",        "30分", "25分", "15分", 1),
    ("チョコデニッシュ",  "45分", "30分", "20分", 2),
    ("バゲット",          "20分", "10分", "30分", 1),
    ("シナモンロール",    "40分", "20分", "25分", 1),
]

COST_DATA = [
    ("食パン（1斤）",     45,  20, 12, 10),
    ("ベーグル（プレーン）",25, 10, 10,  8),
    ("クロワッサン",      30,  45, 15, 12),
    ("カレーパン",        28,  15, 15, 15),
    ("メロンパン",        25,  12, 12,  8),
    ("チョコデニッシュ",  28,  25, 15, 14),
    ("バゲット",          50,  15, 15, 12),
    ("シナモンロール",    30,  30, 15, 12),
]

DATE_FMT  = "YYYY/MM/DD"
YEN_FMT   = '#,##0"円";(#,##0"円");"-"'
YEN_FMT2  = '#,##0;(#,##0);"-"'
PCT_FMT   = '0.0%'
TIME_FMT  = "HH:MM"

ROW_H = 18


# ════════════════════════════════════════════════════════════
# 01_売上入力
# ════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("01_売上入力")
ws1.sheet_properties.tabColor = "70AD47"
ws1.freeze_panes = "A3"

section_title(ws1, 1, 7, "01  売上入力シート　―　日々の販売データを入力してください")
ws1.row_dimensions[1].height = 24

headers = ["日付", "曜日", "時間帯", "商品名", "数量", "単価 (円)", "売上 (円)"]
bgs     = ["2F5597"] * 7
for ci, (h, bg) in enumerate(zip(headers, bgs), 1):
    hdr_cell(ws1, 2, ci, h, bg=bg)
ws1.row_dimensions[2].height = 20

for ri, row in enumerate(SALES_DATA, 3):
    date, dow, slot, name, qty, price = row
    data_cell(ws1, ri, 1, date,  fmt=DATE_FMT)
    data_cell(ws1, ri, 2, dow)
    data_cell(ws1, ri, 3, slot)
    data_cell(ws1, ri, 4, name)
    data_cell(ws1, ri, 5, qty,   fmt=YEN_FMT2)
    data_cell(ws1, ri, 6, price, fmt=YEN_FMT)
    fc = ws1.cell(ri, 7, value=f"=E{ri}*F{ri}")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = YEN_FMT
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws1.row_dimensions[ri].height = ROW_H

# 入力ガイド行（51行まで空白フォーム）
for ri in range(len(SALES_DATA) + 3, 52):
    for ci in range(1, 8):
        c = ws1.cell(ri, ci)
        if ci == 7:
            c.value = f"=IF(E{ri}=\"\",\"\",E{ri}*F{ri})"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
        c.number_format = YEN_FMT if ci in (6, 7) else (DATE_FMT if ci == 1 else YEN_FMT2 if ci == 5 else "General")
        thin = Side(style="thin", color="CCCCCC")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.fill = PatternFill("solid", start_color="F2F2F2")
    ws1.row_dimensions[ri].height = ROW_H

set_col_widths(ws1, {"A":14,"B":6,"C":10,"D":22,"E":8,"F":12,"G":12})


# ════════════════════════════════════════════════════════════
# 02_商品マスタ
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("02_商品マスタ")
ws2.sheet_properties.tabColor = "FFC000"
ws2.freeze_panes = "A3"

section_title(ws2, 1, 8, "02  商品マスタ　―　商品情報・粗利を一元管理")
headers2 = ["商品ID","商品名","カテゴリ","販売価格 (円)","原価 (円)","粗利 (円)","製造時間 (分)","推奨製造数"]
for ci, h in enumerate(headers2, 1):
    hdr_cell(ws2, 2, ci, h, bg="C55A11")

for ri, p in enumerate(PRODUCTS, 3):
    pid, name, cat, price, cost, mtime, rec = p
    data_cell(ws2, ri, 1, pid)
    data_cell(ws2, ri, 2, name)
    data_cell(ws2, ri, 3, cat)
    data_cell(ws2, ri, 4, price, fmt=YEN_FMT)
    data_cell(ws2, ri, 5, cost,  fmt=YEN_FMT)
    # 粗利 = 販売価格 - 原価  (formula)
    fc = ws2.cell(ri, 6, value=f"=D{ri}-E{ri}")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = YEN_FMT
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_cell(ws2, ri, 7, mtime, fmt=YEN_FMT2)
    data_cell(ws2, ri, 8, rec,   fmt=YEN_FMT2)
    ws2.row_dimensions[ri].height = ROW_H

# 粗利率列（I列）
hdr_cell(ws2, 2, 9, "粗利率", bg="C55A11")
for ri in range(3, 3 + len(PRODUCTS)):
    fc = ws2.cell(ri, 9, value=f"=IFERROR(F{ri}/D{ri},0)")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = PCT_FMT
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

set_col_widths(ws2, {"A":10,"B":22,"C":14,"D":14,"E":12,"F":12,"G":14,"H":12,"I":10})


# ════════════════════════════════════════════════════════════
# 03_原価管理
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("03_原価管理")
ws3.sheet_properties.tabColor = "ED7D31"
ws3.freeze_panes = "A3"

section_title(ws3, 1, 6, "03  原価管理　―　材料ごとの原価内訳を管理")
headers3 = ["商品名","小麦 (円)","バター (円)","包材 (円)","その他 (円)","合計原価 (円)"]
for ci, h in enumerate(headers3, 1):
    hdr_cell(ws3, 2, ci, h, bg="843C0C")

for ri, row in enumerate(COST_DATA, 3):
    name, wheat, butter, wrap, other = row
    data_cell(ws3, ri, 1, name)
    data_cell(ws3, ri, 2, wheat,  fmt=YEN_FMT)
    data_cell(ws3, ri, 3, butter, fmt=YEN_FMT)
    data_cell(ws3, ri, 4, wrap,   fmt=YEN_FMT)
    data_cell(ws3, ri, 5, other,  fmt=YEN_FMT)
    fc = ws3.cell(ri, 6, value=f"=SUM(B{ri}:E{ri})")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = YEN_FMT
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws3.row_dimensions[ri].height = ROW_H

set_col_widths(ws3, {"A":22,"B":12,"C":12,"D":12,"E":12,"F":14})


# ════════════════════════════════════════════════════════════
# 04_廃棄入力
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("04_廃棄入力")
ws4.sheet_properties.tabColor = "FF0000"
ws4.freeze_panes = "A3"

section_title(ws4, 1, 5, "04  廃棄入力　―　廃棄した商品を記録")
headers4 = ["日付","商品名","廃棄数","販売価格 (円)","廃棄金額 (円)"]
for ci, h in enumerate(headers4, 1):
    hdr_cell(ws4, 2, ci, h, bg="C00000")

for ri, row in enumerate(WASTE_DATA, 3):
    date, name, qty, price = row
    data_cell(ws4, ri, 1, date,  fmt=DATE_FMT)
    data_cell(ws4, ri, 2, name)
    data_cell(ws4, ri, 3, qty,   fmt=YEN_FMT2)
    data_cell(ws4, ri, 4, price, fmt=YEN_FMT)
    fc = ws4.cell(ri, 5, value=f"=C{ri}*D{ri}")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = YEN_FMT
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws4.row_dimensions[ri].height = ROW_H

for ri in range(len(WASTE_DATA) + 3, 52):
    for ci in range(1, 6):
        c = ws4.cell(ri, ci)
        if ci == 5:
            c.value = f"=IF(C{ri}=\"\",\"\",C{ri}*D{ri})"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
        thin = Side(style="thin", color="CCCCCC")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.fill = PatternFill("solid", start_color="FFF2F2")
    ws4.row_dimensions[ri].height = ROW_H

set_col_widths(ws4, {"A":14,"B":22,"C":10,"D":14,"E":14})


# ════════════════════════════════════════════════════════════
# 05_製造時間
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("05_製造時間")
ws5.sheet_properties.tabColor = "9E480E"
ws5.freeze_panes = "A3"

section_title(ws5, 1, 6, "05  製造時間　―　商品ごとの工程時間と作業負荷を管理")
headers5 = ["商品名","仕込み時間","成形時間","焼成時間","合計時間（分）","作業人数"]
for ci, h in enumerate(headers5, 1):
    hdr_cell(ws5, 2, ci, h, bg="833C00")

# 合計時間を数値（分）で管理
MFG_NUMS = [
    ("食パン（1斤）",     30, 10, 50, 1),
    ("ベーグル（プレーン）",30, 20, 25, 1),
    ("クロワッサン",      60, 40, 20, 2),
    ("カレーパン",        30, 30, 20, 1),
    ("メロンパン",        30, 25, 15, 1),
    ("チョコデニッシュ",  45, 30, 20, 2),
    ("バゲット",          20, 10, 30, 1),
    ("シナモンロール",    40, 20, 25, 1),
]

for ri, row in enumerate(MFG_NUMS, 3):
    name, prep, shape, bake, workers = row
    data_cell(ws5, ri, 1, name)
    data_cell(ws5, ri, 2, prep,    fmt=YEN_FMT2)
    data_cell(ws5, ri, 3, shape,   fmt=YEN_FMT2)
    data_cell(ws5, ri, 4, bake,    fmt=YEN_FMT2)
    fc = ws5.cell(ri, 5, value=f"=SUM(B{ri}:D{ri})")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = YEN_FMT2
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_cell(ws5, ri, 6, workers, fmt=YEN_FMT2)
    ws5.row_dimensions[ri].height = ROW_H

# 利益/時間列（G列）
hdr_cell(ws5, 2, 7, "粗利/時間 (円/時)", bg="833C00")
for ri2, p in enumerate(PRODUCTS, 3):
    pid, pname, cat, price, cost, mtime, rec = p
    fc = ws5.cell(ri2, 7, value=f"=IFERROR((XLOOKUP(A{ri2},'02_商品マスタ'!B:B,'02_商品マスタ'!F:F,0))/(E{ri2}/60),0)")
    fc.font = Font(name=FONT_NAME, size=10, color="000000")
    fc.number_format = YEN_FMT2
    thin = Side(style="thin", color="CCCCCC")
    fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

set_col_widths(ws5, {"A":22,"B":12,"C":12,"D":12,"E":14,"F":12,"G":18})


# ════════════════════════════════════════════════════════════
# 06_人件費
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("06_人件費")
ws6.sheet_properties.tabColor = "7030A0"
ws6.freeze_panes = "A3"

section_title(ws6, 1, 7, "06  人件費　―　シフト・労働時間・人件費を記録")
headers6 = ["日付","スタッフ名","出勤","退勤","時給 (円)","労働時間 (h)","人件費 (円)"]
for ci, h in enumerate(headers6, 1):
    hdr_cell(ws6, 2, ci, h, bg="4B0082")

for ri, row in enumerate(LABOR_DATA, 3):
    date, name, start, end, wage = row
    data_cell(ws6, ri, 1, date,  fmt=DATE_FMT)
    data_cell(ws6, ri, 2, name)
    # 時刻文字列 → time型
    sh, sm = map(int, start.split(":"))
    eh, em = map(int, end.split(":"))
    data_cell(ws6, ri, 3, datetime.time(sh, sm), fmt="HH:MM")
    data_cell(ws6, ri, 4, datetime.time(eh, em), fmt="HH:MM")
    data_cell(ws6, ri, 5, wage, fmt=YEN_FMT)
    fc_h = ws6.cell(ri, 6, value=f"=(D{ri}-C{ri})*24")
    fc_h.font = Font(name=FONT_NAME, size=10, color="000000")
    fc_h.number_format = "0.0"
    thin = Side(style="thin", color="CCCCCC")
    fc_h.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fc_w = ws6.cell(ri, 7, value=f"=F{ri}*E{ri}")
    fc_w.font = Font(name=FONT_NAME, size=10, color="000000")
    fc_w.number_format = YEN_FMT
    fc_w.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws6.row_dimensions[ri].height = ROW_H

for ri in range(len(LABOR_DATA) + 3, 52):
    for ci in range(1, 8):
        c = ws6.cell(ri, ci)
        if ci == 6:
            c.value = f"=IFERROR((D{ri}-C{ri})*24,\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
        elif ci == 7:
            c.value = f"=IFERROR(F{ri}*E{ri},\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT
        thin = Side(style="thin", color="CCCCCC")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.fill = PatternFill("solid", start_color="F5F0FF")
    ws6.row_dimensions[ri].height = ROW_H

set_col_widths(ws6, {"A":14,"B":16,"C":10,"D":10,"E":12,"F":14,"G":14})


# ════════════════════════════════════════════════════════════
# 07_分析ダッシュボード
# ════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("07_分析ダッシュボード")
ws7.sheet_properties.tabColor = "4472C4"
ws7.freeze_panes = "A3"

section_title(ws7, 1, 12, "07  分析ダッシュボード　―　自動集計・可視化")
ws7.row_dimensions[1].height = 26

# ─── KPIブロック（左：A-C列）────────────────────────────
ws7.merge_cells("A3:C3")
c = ws7["A3"]; c.value = "■ KPI サマリ（累計）"
c.font = Font(name=FONT_NAME, bold=True, size=12, color="1F3864")
c.fill = PatternFill("solid", start_color="D9E1F2")

kpi_labels = [
    ("総売上合計 (円)",   "=IFERROR(SUM('01_売上入力'!G3:G200),0)",            YEN_FMT),
    ("総販売個数",        "=IFERROR(SUM('01_売上入力'!E3:E200),0)",            YEN_FMT2),
    ("総廃棄金額 (円)",   "=IFERROR(SUM('04_廃棄入力'!E3:E200),0)",            YEN_FMT),
    ("総人件費 (円)",     "=IFERROR(SUM('06_人件費'!G3:G200),0)",              YEN_FMT),
    ("粗利合計 (円)",     "=IFERROR(B5-B7-B8,0)",                              YEN_FMT),
    ("客単価 (円/人)",    "=IFERROR(B5/SUM(日次サマリ!B3:B200),\"-\")",        YEN_FMT2),
]

hdr_cell(ws7, 4, 1, "指標", bg="2F5597", size=10)
hdr_cell(ws7, 4, 2, "値",   bg="2F5597", size=10)
hdr_cell(ws7, 4, 3, "",     bg="2F5597", size=10)

for ri2, (label, formula, fmt) in enumerate(kpi_labels, 5):
    lc = ws7.cell(ri2, 1, value=label)
    lc.font = Font(name=FONT_NAME, bold=True, size=10)
    lc.fill = PatternFill("solid", start_color="E9EFF7")
    thin = Side(style="thin", color="AAAAAA")
    lc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    lc.alignment = Alignment(vertical="center")
    vc = ws7.cell(ri2, 2, value=formula)
    vc.font = Font(name=FONT_NAME, size=11, color="000000", bold=True)
    vc.number_format = fmt
    vc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    vc.alignment = Alignment(horizontal="right", vertical="center")
    ws7.cell(ri2, 3).fill = PatternFill("solid", start_color="E9EFF7")
    ws7.cell(ri2, 3).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws7.row_dimensions[ri2].height = 22

# ─── 商品別売上ランキング（右：E-I列）──────────────────
ws7.merge_cells("E3:I3")
c = ws7["E3"]; c.value = "■ 商品別売上ランキング"
c.font = Font(name=FONT_NAME, bold=True, size=12, color="1F3864")
c.fill = PatternFill("solid", start_color="D9E1F2")

for ci2, h in enumerate(["商品名","売上 (円)","販売数","廃棄数","粗利 (円)"], 5):
    hdr_cell(ws7, 4, ci2, h, bg="2F5597", size=10)

for ri2, pname in enumerate([p[1] for p in PRODUCTS], 5):
    thin = Side(style="thin", color="CCCCCC")
    nc = ws7.cell(ri2, 5, value=pname)
    nc.font = Font(name=FONT_NAME, size=10)
    nc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci2, (formula, fmt) in enumerate([
        (f"=IFERROR(SUMIF('01_売上入力'!D:D,E{ri2},'01_売上入力'!G:G),0)", YEN_FMT),
        (f"=IFERROR(SUMIF('01_売上入力'!D:D,E{ri2},'01_売上入力'!E:E),0)", YEN_FMT2),
        (f"=IFERROR(SUMIF('04_廃棄入力'!B:B,E{ri2},'04_廃棄入力'!C:C),0)", YEN_FMT2),
        (f"=IFERROR(F{ri2}-XLOOKUP(E{ri2},'02_商品マスタ'!B:B,'02_商品マスタ'!E:E,0)*H{ri2},0)", YEN_FMT),
    ], 6):
        fc = ws7.cell(ri2, ci2, value=formula)
        fc.font = Font(name=FONT_NAME, size=10, color="000000")
        fc.number_format = fmt
        fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws7.row_dimensions[ri2].height = 18

# ─── 時間帯別売上テーブル（A-C列, row 13〜）─────────────
row_t = 13
ws7.merge_cells(f"A{row_t}:C{row_t}")
c = ws7[f"A{row_t}"]; c.value = "■ 時間帯別売上"
c.font = Font(name=FONT_NAME, bold=True, size=12, color="1F3864")
c.fill = PatternFill("solid", start_color="D9E1F2")

time_slots = ["8-9","9-10","10-11","11-12","12-13","13-14","14-15","15-16"]
for ci2, h in enumerate(["時間帯","売上 (円)","販売数"], 1):
    hdr_cell(ws7, row_t+1, ci2, h, bg="2F5597", size=10)

for ri2, slot in enumerate(time_slots, row_t+2):
    thin = Side(style="thin", color="CCCCCC")
    ws7.cell(ri2, 1, value=slot).font = Font(name=FONT_NAME, size=10)
    ws7.cell(ri2, 1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fc_s = ws7.cell(ri2, 2, value=f"=IFERROR(SUMIF('01_売上入力'!C:C,A{ri2},'01_売上入力'!G:G),0)")
    fc_s.font = Font(name=FONT_NAME, size=10, color="000000")
    fc_s.number_format = YEN_FMT
    fc_s.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fc_c = ws7.cell(ri2, 3, value=f"=IFERROR(SUMIF('01_売上入力'!C:C,A{ri2},'01_売上入力'!E:E),0)")
    fc_c.font = Font(name=FONT_NAME, size=10, color="000000")
    fc_c.number_format = YEN_FMT2
    fc_c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws7.row_dimensions[ri2].height = 18

# ─── グラフ：時間帯別売上（棒グラフ）─────────────────
chart_bar = BarChart()
chart_bar.title = "時間帯別売上"
chart_bar.style = 10
chart_bar.y_axis.title = "売上 (円)"
chart_bar.x_axis.title = "時間帯"
chart_bar.width = 20
chart_bar.height = 12

data_ref = Reference(ws7, min_col=2, min_row=row_t+1,
                     max_col=2, max_row=row_t+1+len(time_slots))
cats_ref = Reference(ws7, min_col=1, min_row=row_t+2,
                     max_row=row_t+1+len(time_slots))
chart_bar.add_data(data_ref, titles_from_data=True)
chart_bar.set_categories(cats_ref)
chart_bar.series[0].graphicalProperties.solidFill = "4472C4"
ws7.add_chart(chart_bar, "E13")

# ─── グラフ：商品別売上（横棒グラフ）─────────────────
chart_hbar = BarChart()
chart_hbar.type = "bar"
chart_hbar.title = "商品別売上"
chart_hbar.style = 10
chart_hbar.y_axis.title = "商品"
chart_hbar.x_axis.title = "売上 (円)"
chart_hbar.width = 20
chart_hbar.height = 14

dr2 = Reference(ws7, min_col=6, min_row=4, max_row=4+len(PRODUCTS))
cr2 = Reference(ws7, min_col=5, min_row=5, max_row=4+len(PRODUCTS))
chart_hbar.add_data(dr2, titles_from_data=True)
chart_hbar.set_categories(cr2)
chart_hbar.series[0].graphicalProperties.solidFill = "70AD47"
ws7.add_chart(chart_hbar, "E31")

# ════════════════════════════════════════════════════════════
# 日次サマリ（別シート）
# ════════════════════════════════════════════════════════════
ws_day = wb.create_sheet("日次サマリ")
ws_day.sheet_properties.tabColor = "00B0F0"
ws_day.freeze_panes = "A3"

section_title(ws_day, 1, 6, "日次サマリ　―　日別：売上・客数・客単価・販売個数")
ws_day.row_dimensions[1].height = 24

day_hdrs = ["日付", "客数（手入力）", "日売上 (円)", "販売個数", "客単価 (円/人)", "廃棄金額 (円)"]
day_bgs  = ["1F497D","C55A11","2F5597","2F5597","375623","C00000"]
for ci2, (h, bg) in enumerate(zip(day_hdrs, day_bgs), 1):
    hdr_cell(ws_day, 2, ci2, h, bg=bg)

# 入力ガイド（黄色背景で客数列を強調）
guide = ws_day.cell(2, 2)
guide.fill = PatternFill("solid", start_color="FF8C00")

SAMPLE_DATES = [
    datetime.date(2024, 6, 1),
    datetime.date(2024, 6, 2),
    datetime.date(2024, 6, 3),
]
SAMPLE_CUSTOMERS = [42, 38, 25]

for ri2, (d, cust) in enumerate(zip(SAMPLE_DATES, SAMPLE_CUSTOMERS), 3):
    thin = Side(style="thin", color="CCCCCC")
    dc = ws_day.cell(ri2, 1, value=d)
    dc.number_format = DATE_FMT
    dc.font = Font(name=FONT_NAME, size=10)
    dc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    kc = ws_day.cell(ri2, 2, value=cust)
    kc.font = Font(name=FONT_NAME, size=10, color="0000FF", bold=True)
    kc.number_format = YEN_FMT2
    kc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    kc.fill = PatternFill("solid", start_color="FFFFC0")

    for ci2, (formula, fmt) in enumerate([
        (f"=IFERROR(SUMIFS('01_売上入力'!G:G,'01_売上入力'!A:A,A{ri2}),0)", YEN_FMT),
        (f"=IFERROR(SUMIFS('01_売上入力'!E:E,'01_売上入力'!A:A,A{ri2}),0)", YEN_FMT2),
        (f"=IFERROR(IF(B{ri2}=0,\"-\",C{ri2}/B{ri2}),\"-\")",               YEN_FMT2),
        (f"=IFERROR(SUMIFS('04_廃棄入力'!E:E,'04_廃棄入力'!A:A,A{ri2}),0)", YEN_FMT),
    ], 3):
        fc = ws_day.cell(ri2, ci2, value=formula)
        fc.font = Font(name=FONT_NAME, size=10, color="000000")
        fc.number_format = fmt
        fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws_day.row_dimensions[ri2].height = ROW_H

for ri2 in range(len(SAMPLE_DATES) + 3, 203):
    thin = Side(style="thin", color="CCCCCC")
    for ci2 in range(1, 7):
        c = ws_day.cell(ri2, ci2)
        if ci2 == 3:
            c.value = f"=IFERROR(SUMIFS('01_売上入力'!G:G,'01_売上入力'!A:A,A{ri2}),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT
        elif ci2 == 4:
            c.value = f"=IFERROR(SUMIFS('01_売上入力'!E:E,'01_売上入力'!A:A,A{ri2}),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT2
        elif ci2 == 5:
            c.value = f"=IFERROR(IF(B{ri2}=\"\",\"\",IF(B{ri2}=0,\"-\",C{ri2}/B{ri2})),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT2
        elif ci2 == 6:
            c.value = f"=IFERROR(SUMIFS('04_廃棄入力'!E:E,'04_廃棄入力'!A:A,A{ri2}),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT
        elif ci2 == 2:
            c.fill = PatternFill("solid", start_color="FFFFC0")
            c.font = Font(name=FONT_NAME, size=10, color="0000FF")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.fill = PatternFill("solid", start_color="F5FBFF" if ci2 != 2 else "FFFFC0")
    ws_day.row_dimensions[ri2].height = ROW_H

# 日次グラフ：日別売上（折れ線）
chart_daily = LineChart()
chart_daily.title = "日別売上推移"
chart_daily.style = 10
chart_daily.y_axis.title = "売上 (円)"
chart_daily.x_axis.title = "日付"
chart_daily.width = 22
chart_daily.height = 13

dr_d = Reference(ws_day, min_col=3, min_row=2, max_row=2+len(SAMPLE_DATES))
cr_d = Reference(ws_day, min_col=1, min_row=3, max_row=2+len(SAMPLE_DATES))
chart_daily.add_data(dr_d, titles_from_data=True)
chart_daily.set_categories(cr_d)
chart_daily.series[0].graphicalProperties.line.solidFill = "4472C4"
ws_day.add_chart(chart_daily, "H3")

set_col_widths(ws_day, {"A":14,"B":16,"C":16,"D":12,"E":16,"F":14})

# ════════════════════════════════════════════════════════════
# 月次サマリ（別シート）
# ════════════════════════════════════════════════════════════
ws_mon = wb.create_sheet("月次サマリ")
ws_mon.sheet_properties.tabColor = "7030A0"
ws_mon.freeze_panes = "A3"

section_title(ws_mon, 1, 6, "月次サマリ　―　月別：月商・客数・客単価・廃棄率")
ws_mon.row_dimensions[1].height = 24

mon_hdrs = ["年月", "月商 (円)", "販売個数", "客数", "客単価 (円/人)", "廃棄金額 (円)"]
mon_bgs  = ["1F497D","2F5597","2F5597","C55A11","375623","C00000"]
for ci2, (h, bg) in enumerate(zip(mon_hdrs, mon_bgs), 1):
    hdr_cell(ws_mon, 2, ci2, h, bg=bg)

guide2 = ws_mon.cell(2, 4)
guide2.fill = PatternFill("solid", start_color="FF8C00")

SAMPLE_MONTHS = ["2024/06"]
SAMPLE_MON_CUST = [105]

for ri2, (ym, cust) in enumerate(zip(SAMPLE_MONTHS, SAMPLE_MON_CUST), 3):
    thin = Side(style="thin", color="CCCCCC")
    yc = ws_mon.cell(ri2, 1, value=ym)
    yc.font = Font(name=FONT_NAME, size=10)
    yc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    kc = ws_mon.cell(ri2, 4, value=cust)
    kc.font = Font(name=FONT_NAME, size=10, color="0000FF", bold=True)
    kc.number_format = YEN_FMT2
    kc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    kc.fill = PatternFill("solid", start_color="FFFFC0")

    for ci2, (formula, fmt) in enumerate([
        (f"=IFERROR(SUMPRODUCT((TEXT('01_売上入力'!A3:A200,\"YYYY/MM\")=A{ri2})*('01_売上入力'!G3:G200)),0)", YEN_FMT),
        (f"=IFERROR(SUMPRODUCT((TEXT('01_売上入力'!A3:A200,\"YYYY/MM\")=A{ri2})*('01_売上入力'!E3:E200)),0)", YEN_FMT2),
        (f"=IFERROR(IF(D{ri2}=0,\"-\",B{ri2}/D{ri2}),\"-\")", YEN_FMT2),
        (f"=IFERROR(SUMPRODUCT((TEXT('04_廃棄入力'!A3:A200,\"YYYY/MM\")=A{ri2})*('04_廃棄入力'!E3:E200)),0)", YEN_FMT),
    ], 2):
        fc = ws_mon.cell(ri2, ci2, value=formula)
        fc.font = Font(name=FONT_NAME, size=10, color="000000")
        fc.number_format = fmt
        fc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws_mon.row_dimensions[ri2].height = ROW_H

for ri2 in range(len(SAMPLE_MONTHS) + 3, 50):
    thin = Side(style="thin", color="CCCCCC")
    for ci2 in range(1, 7):
        c = ws_mon.cell(ri2, ci2)
        if ci2 == 2:
            c.value = f"=IFERROR(SUMPRODUCT((TEXT('01_売上入力'!A3:A200,\"YYYY/MM\")=A{ri2})*('01_売上入力'!G3:G200)),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT
        elif ci2 == 3:
            c.value = f"=IFERROR(SUMPRODUCT((TEXT('01_売上入力'!A3:A200,\"YYYY/MM\")=A{ri2})*('01_売上入力'!E3:E200)),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT2
        elif ci2 == 5:
            c.value = f"=IFERROR(IF(D{ri2}=\"\",\"\",IF(D{ri2}=0,\"-\",B{ri2}/D{ri2})),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT2
        elif ci2 == 6:
            c.value = f"=IFERROR(SUMPRODUCT((TEXT('04_廃棄入力'!A3:A200,\"YYYY/MM\")=A{ri2})*('04_廃棄入力'!E3:E200)),\"\")"
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            c.number_format = YEN_FMT
        elif ci2 == 4:
            c.fill = PatternFill("solid", start_color="FFFFC0")
            c.font = Font(name=FONT_NAME, size=10, color="0000FF")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if ci2 != 4:
            c.fill = PatternFill("solid", start_color="F5F0FF")
    ws_mon.row_dimensions[ri2].height = ROW_H

# 月次グラフ：月別売上（棒グラフ）
chart_monthly = BarChart()
chart_monthly.title = "月別売上"
chart_monthly.style = 10
chart_monthly.y_axis.title = "月商 (円)"
chart_monthly.width = 22
chart_monthly.height = 13

dr_m = Reference(ws_mon, min_col=2, min_row=2, max_row=2+len(SAMPLE_MONTHS))
cr_m = Reference(ws_mon, min_col=1, min_row=3, max_row=2+len(SAMPLE_MONTHS))
chart_monthly.add_data(dr_m, titles_from_data=True)
chart_monthly.set_categories(cr_m)
chart_monthly.series[0].graphicalProperties.solidFill = "7030A0"
ws_mon.add_chart(chart_monthly, "H3")

set_col_widths(ws_mon, {"A":12,"B":16,"C":12,"D":12,"E":16,"F":14})

# ws7 シートの列幅
set_col_widths(ws7, {"A":16,"B":16,"C":8,"D":4,"E":22,"F":16,"G":12,"H":12,"I":14})


# ════════════════════════════════════════════════════════════
# 08_改善提案
# ════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("08_改善提案")
ws8.sheet_properties.tabColor = "FF0000"

section_title(ws8, 1, 5, "08  改善提案　―　AIへの貼り付け用・経営課題の整理")

hdr_cell(ws8, 2, 1, "分析区分",   bg="375623")
hdr_cell(ws8, 2, 2, "問題・指標", bg="375623")
hdr_cell(ws8, 2, 3, "考えられる原因", bg="375623")
hdr_cell(ws8, 2, 4, "改善案",     bg="375623")
hdr_cell(ws8, 2, 5, "優先度",     bg="375623")

improvement_rows = [
    ("利益", "=IFERROR(\"粗利率最低商品: \"&INDEX('02_商品マスタ'!B:B,MATCH(MIN('02_商品マスタ'!I:I),'02_商品マスタ'!I:I,0)),\"データなし\")",
     "原価高騰 / 販売価格が低い",  "値上げ検討 or 原価見直し", "高"),
    ("廃棄", "=IFERROR(\"廃棄金額合計: \"&TEXT(SUM('04_廃棄入力'!E:E),\"#,##0\")&\"円\",\"データなし\")",
     "製造数が需要を上回っている",  "製造数を販売実績に合わせて削減", "高"),
    ("製造", "=IFERROR(\"最も製造時間がかかる商品: \"&INDEX('05_製造時間'!A:A,MATCH(MAX('05_製造時間'!E:E),'05_製造時間'!E:E,0)),\"データなし\")",
     "工程が複雑 / 人手不足",       "製造工程の見直し or 販売終了検討", "中"),
    ("人件費", "=IFERROR(\"人件費合計: \"&TEXT(SUM('06_人件費'!G:G),\"#,##0\")&\"円\",\"データなし\")",
     "シフト過多 / 残業",           "売上ピーク時間帯に集中したシフト設計", "中"),
    ("時間帯", "ピーク時間帯の売上集中確認",
     "焼成タイミングとピーク時間がズレている", "時間帯別データを元に焼成スケジュールを調整", "高"),
]

for ri2, row in enumerate(improvement_rows, 3):
    cat, issue, cause, plan, pri = row
    thin = Side(style="thin", color="CCCCCC")
    for ci2, val in enumerate([cat, issue, cause, plan, pri], 1):
        c = ws8.cell(ri2, ci2, value=val)
        c.font = Font(name=FONT_NAME, size=10, color="000000")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if ci2 == 5:
            c.fill = PatternFill("solid", start_color="FFFF00" if val == "高" else "E2EFDA")
    ws8.row_dimensions[ri2].height = 40

# Claude貼り付け用の説明
row_note = len(improvement_rows) + 5
ws8.merge_cells(f"A{row_note}:E{row_note}")
note = ws8[f"A{row_note}"]
note.value = (
    "【Claude / ChatGPT への貼り付け方】\n"
    "このシートの内容をコピーして以下のプロンプトと一緒に貼り付けてください：\n"
    "「以下はパン屋の経営データ分析です。改善策を具体的に提案してください。」"
)
note.font = Font(name=FONT_NAME, size=10, color="1F3864", italic=True)
note.alignment = Alignment(wrap_text=True, vertical="top")
note.fill = PatternFill("solid", start_color="EBF3FB")
ws8.row_dimensions[row_note].height = 60

set_col_widths(ws8, {"A":14,"B":40,"C":28,"D":28,"E":10})

# ════════════════════════════════════════════════════════════
# 保存
# ════════════════════════════════════════════════════════════
out_path = r"C:\Users\81908\OneDrive\デスクトップ\ニチゲツ会社\パン屋経営分析テンプレートv2.xlsx"
wb.save(out_path)
print(f"保存完了: {out_path}")
